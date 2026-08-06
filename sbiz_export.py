"""
generar_sbiz.py — Completa las 3 plantillas de SBIZ con datos reales de
Contabilium Uruguay.

Decisiones tomadas con Mariano (05-ago-2026):
  * Catálogo: solo marcas propias (SUPRABOND + BULIT + SOMERSET).
  * Categoría / Sub-categoría: se heredan del catálogo argentino que trae
    la plantilla, cruzando por código normalizado. Los que no matchean
    llevan una propuesta manual, marcada para revisión.
  * Listas de precios: base + INTERIOR 2024 (+5%) + GRANDES SUPERFICIES,
    precios SIN IVA.
  * Clientes: los 1.123. Los que no tienen lista en Contabilium van a la
    lista base.
  * Ruta: el IdUsuarioAdicional de Contabilium.

Lo que Contabilium NO tiene y queda pendiente de que lo complete Mariano
está agrupado en el bloque PENDIENTES de abajo.
"""

from __future__ import annotations

import re
import shutil
import sys
import tomllib
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

REPO = Path(__file__).resolve().parent
ASSETS = REPO / "assets"
SALIDA = ASSETS / "SBIZ - GSU"

sys.path.insert(0, str(REPO))


# =====================================================================
# Datos que no existen en Contabilium — los pasó Mariano (06-ago-2026)
# =====================================================================
# Celulares sin espacios ni guiones ni código de país, como pide el
# comentario de la columna en la plantilla.
TELEFONO_VENDEDOR: dict[int, str] = {
    237: "095210365",   # Mario Fleitas
    346: "095970830",   # Arturo Trias
    366: "098514310",   # Marcelo Riso
    666: "095067705",   # Néstor de los Santos
}
JEFE: str = "Ernesto Abreu"       # el mismo para los cuatro vendedores
PEDIDO_MINIMO: float | str = ""   # en Uruguay no hay pedido mínimo
SUCURSAL: int = 1                 # una sola sucursal en Uruguay
NOMBRE_DUENIO: str = ""           # no se releva en ningún sistema

# Poner en True para usar la zona Z-xx de Contabilium como Ruta en lugar
# del IdUsuarioAdicional.
RUTA_POR_ZONA: bool = False


# =====================================================================
# Constantes de Contabilium
# =====================================================================
RUBROS = {
    2009: "AQUALAF", 3883: "Bosch", 1594: "BULIT", 1598: "COMPRAS",
    3882: "DREMEL", 4906: "General", 1596: "INSUMOS", 1597: "MARKETING",
    2010: "Peirano", 1819: "SOMERSET", 1595: "SUPRABOND",
}
MARCAS_PROPIAS = {"SUPRABOND", "BULIT", "SOMERSET"}

NOMBRE_VENDEDOR = {
    237: "Mario Fleitas", 346: "Arturo Trias",
    366: "Marcelo Riso", 666: "Néstor de los Santos",
}

# Numeración de listas para SBIZ. TIENE QUE COINCIDIR con la que SBIZ
# configure del otro lado; hoy es una convención nuestra.
LISTA_BASE, LISTA_INTERIOR, LISTA_GRANDES = 1, 2, 3
ID_LISTA_CONTABILIUM = {91: LISTA_INTERIOR, 100: LISTA_GRANDES, 67: LISTA_BASE}
IVA = 1.22

# Reparación de los acentos rotos en los nombres de Contabilium. Cada
# '?' del dato original es una sola vocal acentuada o un '°'.
ACENTOS = {
    "INSTANT?NEA": "INSTANTÁNEA", "ACR?LICO": "ACRÍLICO", "S?MIL": "SÍMIL",
    "N?": "N°", "AUTOM?TICO": "AUTOMÁTICO", "F?RMULA": "FÓRMULA",
    "CA?O": "CAÑO", "CA?OS": "CAÑOS", "MARR?N": "MARRÓN",
    "MURCI?LAGOS": "MURCIÉLAGOS", "ZINGUER?A": "ZINGUERÍA",
    "CONSTRUCCI?N": "CONSTRUCCIÓN", "AVIACI?N": "AVIACIÓN",
    "L?QUIDA": "LÍQUIDA", "Z?CALO": "ZÓCALO",
}

# Categoría / Sub-categoría propuesta para los productos uruguayos que no
# existen en el catálogo argentino. Van marcados en la hoja de revisión.
PROPUESTA = {
    "NSS 10": ("Adhesivos", "Adhesivos Instantáneos"),
    "CAB": ("Engrampadoras - Grampas y Clavos", "Engrampadoras"),
    "CAB PRO": ("Engrampadoras - Grampas y Clavos", "Engrampadoras"),
    "CAD 1/CAD 1/4 14 1": ("Candados y Seguridad", "Seguridad Cadenas"),
    "CDB T DISCO": ("Candados y Seguridad", "Seguridad Trabadiscos"),
    "CDB B 25": ("Candados y Seguridad", "Candados con Llave"),
    "COM SBD TR C": ("Adhesivos", "Adhesivos De Contacto"),
    "CTA MCA7F 8": ("Herramientas de Medición", "Cintas métricas"),
    "C CTR SBD C": ("Herramientas de Corte", "Cutters"),
    "GAS 125 E": ("Selladores y Pistolas Aplicadoras", "Selladores Sellarosca Gas"),
    "SPS P49": ("Adhesivos", "Adhesivos Especiales"),
    "JD S8 PZA BY": ("Jardinería", "Jardinería Herramientas de corte"),
    "ZNO B I": ("Burletes y Zócalos", "Zócalo Autoadhesivo"),
    "ZNO M I": ("Burletes y Zócalos", "Zócalo Autoadhesivo"),
    "PZA FZA 6.5": ("Herramientas de Mano", "Pinza y Alicate"),
}
# Reglas por prefijo, para lo que cae en familias enteras.
PROPUESTA_PREFIJO = [
    ("PZA ", ("Herramientas de Mano", "Pinza y Alicate")),
    ("DES ", ("Herramientas de Mano", "Destornilladores")),
    ("PBD ", ("Adhesivos", "Adhesivos De Contacto")),
    ("SGC ", ("Control de Plagas y Repelentes   ", "Somerset Cucarachas")),
]

norm = lambda s: re.sub(r"[^A-Z0-9]", "", str(s or "").upper())


def reparar(nombre: str) -> str:
    """Devuelve el nombre con los acentos rotos reparados."""
    if "?" not in nombre:
        return nombre
    out = nombre
    for malo, bueno in sorted(ACENTOS.items(), key=lambda kv: -len(kv[0])):
        out = out.replace(malo, bueno)
    return out


def categoria_propuesta(sku: str) -> tuple[str, str] | None:
    if sku in PROPUESTA:
        return PROPUESTA[sku]
    for pref, cat in PROPUESTA_PREFIJO:
        if sku.startswith(pref):
            return cat
    return None


def telefono_limpio(raw: str) -> str:
    """Un solo número, sin espacios ni guiones ni código de país.

    Contabilium guarda hasta dos números separados por '/'. Se prefiere
    el celular (empieza en 09) porque es el que sirve para contactar al
    cliente; si no hay, se usa el primero.
    """
    if not raw:
        return ""
    partes = [re.sub(r"\D", "", p) for p in re.split(r"[/,;]", raw)]
    partes = [p[3:] if p.startswith("598") and len(p) > 9 else p for p in partes]
    partes = [p for p in partes if p]
    if not partes:
        return ""
    for p in partes:
        if p.startswith("09"):
            return p
    return partes[0]


# =====================================================================
# Carga de datos
# =====================================================================
def cargar_contabilium() -> tuple[list[dict], list[dict], dict[int, dict]]:
    """Conceptos, clientes y las 3 listas de precio, desde la API."""
    import api_loader as A

    sec = tomllib.load(open(REPO / ".streamlit/secrets.toml", "rb"))
    s = A.obtener_token(sec["contabilium_client_id"], sec["contabilium_client_secret"])

    s, conceptos = A.api_paginate(s, "/api/conceptos/search")
    s, clientes = A.api_paginate(s, "/api/clientes/search")

    listas: dict[int, dict] = {}
    for lid in (100, 91):
        items, page = [], 1
        while True:
            s, r = A.api_get(s, f"/api/listasDePrecio/getById?id={lid}&page={page}&pageSize=50")
            items += r["Items"]
            if page >= r["TotalPage"] or not r["Items"]:
                break
            page += 1
        listas[lid] = {i["Codigo"]: i for i in items}
    return conceptos, clientes, listas


def cargar_catalogo_ar() -> tuple[dict[str, list], dict[str, str]]:
    """Filas del catálogo argentino por código normalizado + familias."""
    wb = openpyxl.load_workbook(ASSETS / "Plantilla Catálogo - SBIZ.xlsx", data_only=True)
    ar: dict[str, list] = {}
    for r in wb["Catálogo 22.06"].iter_rows(min_row=2, values_only=True):
        if r[5]:
            ar.setdefault(norm(r[5]), []).append(r)
    # 3 productos viven solo en la hoja de orden; los sumamos con los
    # campos que esa hoja sí tiene.
    for r in wb["Orden de Productos"].iter_rows(min_row=2, values_only=True):
        k = norm(r[0])
        if r[0] and k not in ar:
            ar[k] = [(r[0], r[2], r[3], r[1], r[4], r[0], r[5], r[5],
                      1, None, None, None, None, None, None, None)]

    wb2 = openpyxl.load_workbook(ASSETS / "sku_familia_subgrupo.xlsx", data_only=True)
    fam = {norm(r[0]): str(r[1]).strip() for r in wb2.active.iter_rows(min_row=2, values_only=True) if r[0] and r[1]}
    return ar, fam


# =====================================================================
# Construcción de las filas
# =====================================================================
def construir_catalogo(conceptos, ar, fam):
    """Una fila por producto-imagen, como hace el catálogo argentino."""
    productos = [
        c for c in conceptos
        if c.get("Tipo") == "Producto"
        and RUBROS.get(int(c.get("IdRubro") or 0)) in MARCAS_PROPIAS
    ]
    filas, revision, correcciones = [], [], []
    for c in productos:
        sku = (c.get("Codigo") or "").strip()
        k = norm(sku)
        marca = RUBROS[int(c["IdRubro"])]
        nombre_raw = (c.get("Nombre") or "").strip()
        nombre = reparar(nombre_raw)
        if nombre != nombre_raw:
            correcciones.append((sku, nombre_raw, nombre))
        desc = (c.get("Descripcion") or "").strip() or nombre

        base_ar = ar.get(k)
        if base_ar:
            cat, subcat = base_ar[0][1], base_ar[0][2]
            agrup = base_ar[0][4] or fam.get(k, "")
            origen = "catálogo AR"
        else:
            prop = categoria_propuesta(sku)
            cat, subcat = prop if prop else ("", "")
            agrup = fam.get(k, "")
            origen = "propuesta (revisar)" if prop else "SIN CATEGORÍA"
            revision.append((sku, nombre, marca, cat, subcat, origen))

        # Una fila por imagen heredada; si no hay match, una sola fila.
        variantes = base_ar or [None]
        for v in variantes:
            filas.append([
                sku,                        # A  (repite el código)
                cat, subcat, marca, agrup, sku, nombre, desc,
                (v[8] if v else 1) or 1,    # Unidades
                None,                       # Presentación Defecto
                v[10] if v else None,       # EAN
                v[11] if v else None,       # imagen
                v[12] if v else None,       # Envase
                v[13] if v else None,       # Inner Box
                v[14] if v else None,       # Presentación
                None,                       # '/'
            ])
    filas.sort(key=lambda f: (str(f[1]), str(f[2]), str(f[6]), str(f[5]), str(f[11] or "")))
    return filas, revision, correcciones


def construir_clientes(clientes):
    filas, sin_vendedor = [], []
    for c in clientes:
        vid = c.get("IdUsuarioAdicional")
        vendedor = NOMBRE_VENDEDOR.get(vid, "")
        zona = (c.get("PisoDepto") or "").strip()
        if RUTA_POR_ZONA:
            ruta = zona if re.fullmatch(r"Z-\d+", zona) else ""
        else:
            ruta = str(vid) if vid else ""
        if not vendedor:
            sin_vendedor.append((c.get("Codigo"), c.get("RazonSocial"), c.get("Ciudad"), vid))

        lista = ID_LISTA_CONTABILIUM.get(c.get("IdListaPrecio"), LISTA_BASE)
        negocio = (c.get("NombreFantasia") or "").strip() or (c.get("RazonSocial") or "").strip()
        dir3 = (c.get("Ciudad") or "").strip() or (c.get("Provincia") or "").strip()
        # PisoDepto guarda la zona de reparto, no un piso: no va en la
        # dirección.
        filas.append([
            SUCURSAL,
            str(c.get("Codigo") or "").strip(),
            vendedor,
            ruta,
            TELEFONO_VENDEDOR.get(vid, ""),
            JEFE,
            negocio,
            NOMBRE_DUENIO,
            lista,
            (c.get("Domicilio") or "").strip(),
            None,
            dir3,
            str(c.get("Cp") or "").strip(),
            telefono_limpio(c.get("Telefono") or ""),
            PEDIDO_MINIMO,
            (c.get("Email") or "").strip(),
        ])
    return filas, sin_vendedor


def construir_listas(conceptos, listas, skus_catalogo):
    """(sku, precio sin IVA) por cada una de las 3 listas."""
    base = {
        (c.get("Codigo") or "").strip(): c
        for c in conceptos if c.get("Tipo") == "Producto"
    }
    out, sin_precio = {}, []
    for etiqueta, fuente in (
        ("base", None), ("interior", listas[91]), ("grandes", listas[100]),
    ):
        filas = []
        for sku in skus_catalogo:
            if fuente is None:
                pf = float(base.get(sku, {}).get("PrecioFinal") or 0)
            else:
                pf = float(fuente.get(sku, {}).get("PrecioFinal") or 0)
            precio = round(pf / IVA, 2)
            if precio <= 0:
                sin_precio.append((etiqueta, sku))
            filas.append((sku, precio))
        out[etiqueta] = filas
    return out, sin_precio


# =====================================================================
# Escritura de los archivos
# =====================================================================
def limpiar(ws, primera_fila: int):
    if ws.max_row >= primera_fila:
        ws.delete_rows(primera_fila, ws.max_row - primera_fila + 1)


def escribir_catalogo(filas):
    dst = SALIDA / "Catálogo - SBIZ - GSU.xlsx"
    shutil.copy(ASSETS / "Plantilla Catálogo - SBIZ.xlsx", dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["Catálogo 22.06"]
    limpiar(ws, 2)
    for f in filas:
        ws.append(f)

    # Hoja de orden: un renglón por producto, en el mismo orden.
    wo = wb["Orden de Productos"]
    limpiar(wo, 2)
    vistos = set()
    for f in filas:
        if f[5] in vistos:
            continue
        vistos.add(f[5])
        wo.append([f[5], f[3], f[1], f[2], f[4], f[6]])

    # La hoja Granularidad es de atributos que Contabilium no tiene.
    limpiar(wb["Granularidad"], 2)
    wb.save(dst)
    return dst, len(vistos)


def escribir_clientes(filas):
    dst = SALIDA / "Clientes - SBIZ - GSU.xlsx"
    shutil.copy(ASSETS / "Plantilla Clientes - SBIZ.xlsx", dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb["Hoja1"]
    limpiar(ws, 2)
    for f in filas:
        ws.append(f)
    # El código de cliente, el CP y los dos teléfonos llevan ceros a la
    # izquierda: si Excel los toma como número, se los come.
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for idx in (1, 4, 12, 13):     # Código, tel vendedor, CP, tel cliente
            row[idx].number_format = "@"
            if row[idx].value is not None:
                row[idx].value = str(row[idx].value)
    wb.save(dst)
    return dst


def escribir_listas(listas_filas, skus_catalogo):
    dst = SALIDA / "Listas de Precios - SBIZ - GSU.xlsm"
    shutil.copy(ASSETS / "Plantilla Listas de Precios - SBIZ.xlsm", dst)
    wb = openpyxl.load_workbook(dst, keep_vba=True)

    nombres = {
        "base": (f"Lista {LISTA_BASE}", "Lista 1"),
        "interior": (f"Lista {LISTA_INTERIOR}", "Lista 5"),
        "grandes": (f"Lista {LISTA_GRANDES}", "Lista 15"),
    }
    for sobrante in ("Lista 20", "Lista 30"):
        del wb[sobrante]

    for etiqueta, (titulo, hoja_origen) in nombres.items():
        ws = wb[hoja_origen]
        ws.title = titulo
        limpiar(ws, 2)
        for i, (sku, precio) in enumerate(listas_filas[etiqueta], start=2):
            ws.cell(i, 1, sku)
            ws.cell(i, 2, precio)
            ws.cell(i, 3, 1)
            ws.cell(i, 4, f'=IFERROR(VLOOKUP(A{i},Catálogo!A:B,2,FALSE),"No está en Catálogo")')

    wc = wb["Catálogo"]
    limpiar(wc, 2)
    for sku in skus_catalogo:
        wc.append([sku, "En catálogo ok"])

    # Orden de hojas: las 3 listas y después el catálogo de control.
    wb._sheets = [wb[f"Lista {n}"] for n in (LISTA_BASE, LISTA_INTERIOR, LISTA_GRANDES)] + [wc]
    wb.save(dst)
    return dst


def escribir_revision(revision, correcciones, sin_vendedor, sin_precio, filas_cat):
    dst = SALIDA / "_REVISAR - SBIZ - GSU.xlsx"
    wb = openpyxl.Workbook()
    neg = Font(bold=True)

    ws = wb.active
    ws.title = "Categorías a revisar"
    ws.append(["SKU", "Nombre", "Marca", "Categoría propuesta", "Sub-categoría propuesta", "Origen"])
    for f in revision:
        ws.append(list(f))

    ws = wb.create_sheet("Nombres corregidos")
    ws.append(["SKU", "Como está en Contabilium", "Como queda en SBIZ"])
    for f in correcciones:
        ws.append(list(f))

    ws = wb.create_sheet("Clientes sin vendedor")
    ws.append(["Código", "Razón social", "Ciudad", "IdUsuarioAdicional"])
    for f in sin_vendedor:
        ws.append(list(f))

    ws = wb.create_sheet("Productos sin precio")
    ws.append(["Lista", "SKU"])
    for f in sin_precio:
        ws.append(list(f))

    ws = wb.create_sheet("Sin imagen")
    ws.append(["SKU", "Nombre"])
    vistos = set()
    for f in filas_cat:
        if f[11] is None and f[5] not in vistos:
            vistos.add(f[5])
            ws.append([f[5], f[6]])

    for hoja in wb.worksheets:
        for c in hoja[1]:
            c.font = neg
        hoja.freeze_panes = "A2"
    wb.save(dst)
    return dst


def main():
    SALIDA.mkdir(parents=True, exist_ok=True)
    print("Bajando Contabilium…")
    conceptos, clientes, listas = cargar_contabilium()
    print(f"  {len(conceptos)} conceptos, {len(clientes)} clientes")

    ar, fam = cargar_catalogo_ar()
    filas_cat, revision, correcciones = construir_catalogo(conceptos, ar, fam)
    skus = list(dict.fromkeys(f[5] for f in filas_cat))
    filas_cli, sin_vendedor = construir_clientes(clientes)
    listas_filas, sin_precio = construir_listas(conceptos, listas, skus)

    p1, n_prod = escribir_catalogo(filas_cat)
    p2 = escribir_clientes(filas_cli)
    p3 = escribir_listas(listas_filas, skus)
    p4 = escribir_revision(revision, correcciones, sin_vendedor, sin_precio, filas_cat)

    print(f"\nCatálogo:  {n_prod} productos en {len(filas_cat)} filas → {p1.name}")
    print(f"Clientes:  {len(filas_cli)} → {p2.name}")
    print(f"Listas:    {len(skus)} productos × 3 listas → {p3.name}")
    print(f"Revisar:   {len(revision)} categorías, {len(correcciones)} nombres, "
          f"{len(sin_vendedor)} clientes sin vendedor, {len(sin_precio)} sin precio → {p4.name}")


if __name__ == "__main__":
    main()
