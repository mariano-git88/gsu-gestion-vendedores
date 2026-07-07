"""
rendicion.py — Lógica de la Rendición de Cobranzas GSU (Fase 1: SIMULADOR).

Automatiza el análisis de la "Planilla de Rendición de Cobranzas" que
arman los vendedores: por cada fila (un recibo) busca la/s factura/s en
Contabilium, calcula la Nota de Crédito por descuento comercial (10%) y
el cobro esperado, y compara contra lo que el vendedor rindió.

IMPORTANTE — esta Fase 1 es SOLO LECTURA. No crea NC, ni recibos, ni
imputa nada en Contabilium. Produce un reporte OK / REVISAR para que
Operaciones valide la lógica con riesgo cero antes de automatizar la
escritura (Fase 2). Ver decisión de diseño 2026-07-01.

Flujo de negocio (definido por Valeria Falero, Operaciones):
  - Factura original = 100%.
  - Si aplica descuento comercial: NC = 10%, cobro esperado = 90%.
  - Si es pago total (sin descuento): cobro esperado = 100%, sin NC.
  - Se compara el cobro real (efectivo + cheque de la planilla) contra el
    esperado con una tolerancia de ±$TOLERANCIA. Dentro de tolerancia →
    OK (automatizable). Fuera → REVISAR manual.

Cómo se determina el 10% (decisión 2026-07-01: "columna dedicada + cotejo
por monto"):
  1. Columna "Descuento" del template (10% / No). Es la fuente primaria.
  2. Si la columna está vacía, se INFIERE por el monto (cobrado ≈ 90% →
     10%; cobrado ≈ 100% → pago total) y se marca `descuento_asumido`
     para que Operaciones lo revise.

El módulo es puro y testeable: `analizar()` recibe un índice de facturas
ya armado (dict), así que la lógica se prueba sin tocar la API. La capa
de red (`construir_indice_facturas`, `obtener_saldo_factura`) se apoya en
`api_loader` (throttle + retry + refresh de token ya resueltos ahí).
"""

from __future__ import annotations

import io
import re
import unicodedata
from dataclasses import dataclass, field
from datetime import datetime

import openpyxl
import pandas as pd

import api_loader

# =====================================================================
# Constantes de negocio
# =====================================================================

TASA_DESCUENTO = 0.10        # descuento comercial estándar (10%)
TOLERANCIA_DEFAULT = 10.0    # ±$ para considerar que el cobro "cuadra"

# TipoFc de Contabilium que NO son facturas cobrables sino notas de crédito
# (reducen deuda). Si un vendedor pone uno de estos números en la columna
# "Nro Factura", es un error: no se cobra contra una NC. Se detecta por tipo
# (señal robusta) y se manda a REVISAR. Espejo de TIPOS_NEGATIVOS de api_loader.
TIPOS_NOTA_CREDITO = frozenset({"NCF", "NCT", "NCE"})

ESTADO_OK = "OK"
ESTADO_REVISAR = "REVISAR"

# Mapa de encabezado normalizado → clave interna. Se matchea por texto
# normalizado (sin acentos, minúsculas) para que aguante que Valeria
# reordene columnas o cambie mayúsculas/puntuación. El orden importa:
# las claves más específicas ("nro cheque") deben resolverse antes que
# las genéricas ("cheque"), por eso se chequea con `in` en ese orden.
#
# NOTA sobre "Nº": al normalizar, el ordinal "º" pierde su decomposición y
# "Nº" queda como "no" (ej. "Nº Cheque" → "no cheque"). Por eso cada alias
# con número incluye también la variante "no ...".
_COLUMNAS = [
    # (clave_interna, [alias normalizados, del más específico al más laxo])
    ("fecha", ["fecha"]),
    ("nro_recibo", ["nro recibo", "no recibo", "numero recibo", "recibo"]),
    ("nro_cliente", ["nro cliente", "no cliente", "numero cliente", "cliente"]),
    ("nro_factura", ["nro factura", "no factura", "numero factura", "factura"]),
    ("descuento", ["descuento", "dto", "10%"]),
    ("nro_cheque", ["nro cheque", "no cheque", "n cheque", "numero cheque"]),
    ("observaciones", ["observaciones", "observacion", "obs"]),
    ("efectivo", ["cobro efectivo", "efectivo"]),
    ("cheque", ["cobro cheque", "cheque"]),
    ("total_recibo", ["total recibo", "total"]),
]


# =====================================================================
# Helpers de normalización
# =====================================================================

def _norm(texto) -> str:
    """Minúsculas, sin acentos, espacios colapsados. Para matchear headers."""
    s = str(texto or "").strip().lower()
    s = "".join(
        c for c in unicodedata.normalize("NFKD", s) if not unicodedata.combining(c)
    )
    s = s.replace(".", " ").replace("-", " ").replace("_", " ")
    return re.sub(r"\s+", " ", s).strip()


def _norm_factura(numero) -> str:
    """Clave de match fuerte para un Nº de factura: solo alfanuméricos.

    'A-00033123' → 'a00033123'. Tolera espacios/puntos/guiones que
    difieran entre la planilla y el campo `Numero` de Contabilium.
    """
    return re.sub(r"[^0-9a-z]", "", str(numero or "").lower())


def _sufijo_num(numero) -> str:
    """Último grupo de dígitos sin ceros a la izquierda: match débil de respaldo.

    'A-00033123' → '33123'. Se usa cuando el match fuerte falla (ej. la
    serie/letra difiere de formato), aceptándolo solo si es inequívoco.
    """
    grupos = re.findall(r"\d+", str(numero or ""))
    return grupos[-1].lstrip("0") if grupos else ""


def _parse_flag_descuento(valor) -> bool | None:
    """Interpreta la columna 'Descuento'. Devuelve True (aplica 10%),
    False (pago total, sin NC) o None (no especificado → inferir por monto).
    """
    if valor is None or (isinstance(valor, str) and not valor.strip()):
        return None
    # Numérico: 0.1 / 10 → aplica; 0 → no aplica.
    if isinstance(valor, (int, float)):
        return abs(float(valor)) > 1e-9
    s = _norm(valor)
    if s in ("no", "n", "0", "0%", "sin", "sin descuento", "pago total", "total", "100%"):
        return False
    if s in ("si", "s", "10", "10%", "0 1", "01", "descuento", "dto", "aplica"):
        return True
    # "10% a la 33248" y similares → contiene 10 o %/descuento ⇒ aplica.
    if "10" in s or "%" in s or "descuento" in s or "dto" in s:
        return True
    return None


def _to_float(valor) -> float:
    """Monto de una celda de la planilla → float. Usa el parser UY de
    api_loader por si viene como string '1.000,00'."""
    if valor is None:
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    return api_loader.parse_monto_uy(valor)


# =====================================================================
# Modelos
# =====================================================================

@dataclass
class FilaCobranza:
    """Una fila de la planilla (= un recibo a registrar)."""
    fila_excel: int
    fecha: datetime | None
    nro_recibo: str
    nro_cliente: str
    facturas: list[str]          # 1+ (multi-factura viene "A-1/A-2")
    descuento_flag: bool | None  # None = no especificado en la columna
    nro_cheque: str
    observaciones: str
    efectivo: float
    cheque: float
    total_recibo: float

    @property
    def es_multifactura(self) -> bool:
        return len(self.facturas) > 1

    @property
    def cobrado(self) -> float:
        """Plata real rendida. Si 'Total Recibo' vino cargado se usa ese;
        si no, se reconstruye como efectivo + cheque."""
        return self.total_recibo if self.total_recibo else (self.efectivo + self.cheque)


@dataclass
class ResultadoFila:
    """Análisis de una fila: qué debería pasar y si cuadra."""
    fila: FilaCobranza
    estado: str
    motivos: list[str] = field(default_factory=list)
    total_factura: float = 0.0
    saldo_factura: float = 0.0
    descuento_aplica: bool = False
    descuento_asumido: bool = False   # True si se infirió por monto
    es_pago_parcial: bool = False     # entrega / pago parcial: no lleva NC
    nc_esperada: float = 0.0
    cobro_esperado: float = 0.0
    diferencia: float = 0.0           # cobrado - cobro_esperado
    facturas_detalle: list[dict] = field(default_factory=list)

    @property
    def es_ejecutable(self) -> bool:
        """True si la Fase 2 puede ejecutar esta fila automáticamente.

        Solo el caso limpio: UNA factura encontrada (no NC), no pago
        parcial. Multi-factura (distribución desconocida) y entregas
        quedan fuera aunque Valeria las apruebe — se cargan a mano.
        """
        if self.es_pago_parcial:
            return False
        if len(self.facturas_detalle) != 1:
            return False
        comp = self.facturas_detalle[0].get("comprobante")
        if not comp:
            return False
        return str(comp.get("tipo") or "").upper() not in TIPOS_NOTA_CREDITO

    def params_ejecucion(self) -> dict | None:
        """Datos que necesita `rendicion_ejecutor` para esta fila, o None
        si no es ejecutable."""
        if not self.es_ejecutable:
            return None
        comp = self.facturas_detalle[0]["comprobante"]
        return {
            "id_factura": comp["id"],
            "numero_factura": comp["numero"],
            "aplica_nc": self.descuento_aplica,
            "cobro_efectivo": self.fila.efectivo,
            "cobro_cheque": self.fila.cheque,
            "nro_cheque": self.fila.nro_cheque,
        }


# =====================================================================
# Lectura de la planilla
# =====================================================================

def _detectar_header(ws) -> tuple[int, dict[str, int]]:
    """Encuentra la fila de encabezados y mapea clave_interna → nº columna.

    Busca en las primeras 15 filas la que contenga 'nro factura'. Devuelve
    (fila_header, {clave: col}). Levanta ValueError si no la encuentra.
    """
    for r in range(1, min(ws.max_row, 15) + 1):
        celdas = {c: _norm(ws.cell(row=r, column=c).value)
                  for c in range(1, ws.max_column + 1)}
        if any("factura" in v for v in celdas.values()):
            mapa: dict[str, int] = {}
            for col, texto in celdas.items():
                if not texto:
                    continue
                for clave, alias in _COLUMNAS:
                    if clave in mapa:
                        continue
                    if any(a in texto for a in alias):
                        mapa[clave] = col
                        break
            return r, mapa
    raise ValueError(
        "No se encontró la fila de encabezados (se esperaba una columna "
        "'Nro Factura'). ¿Es la planilla de Rendición de Cobranzas?"
    )


def leer_planilla(
    fuente, hoja: str | None = None
) -> tuple[list[FilaCobranza], list[dict]]:
    """Lee la planilla de rendición.

    Devuelve (filas, descartadas):
      - `filas`: filas válidas (tienen al menos un Nº de factura).
      - `descartadas`: filas que tenían algún dato (plata o cliente) pero
        NO un Nº de factura, con el motivo. Se reportan aparte para que
        nada real quede silenciosamente ignorado (ej. el vendedor olvidó
        la factura), sin ensuciar el análisis con filas de relleno.

    `fuente` puede ser un path, un file-like o bytes (el uploader de
    Streamlit entrega bytes). `hoja` opcional; por defecto la primera con
    encabezados válidos (típicamente 'Cobranzas').
    """
    if isinstance(fuente, bytes):
        fuente = io.BytesIO(fuente)
    wb = openpyxl.load_workbook(fuente, data_only=True)
    ws = wb[hoja] if hoja else wb[wb.sheetnames[0]]

    fila_header, mapa = _detectar_header(ws)
    if "nro_factura" not in mapa:
        raise ValueError("La planilla no tiene columna de Nº de Factura.")

    def val(row, clave):
        col = mapa.get(clave)
        return ws.cell(row=row, column=col).value if col else None

    filas: list[FilaCobranza] = []
    descartadas: list[dict] = []
    for r in range(fila_header + 1, ws.max_row + 1):
        crudo_factura = val(r, "nro_factura")
        efectivo = _to_float(val(r, "efectivo"))
        cheque = _to_float(val(r, "cheque"))
        total = _to_float(val(r, "total_recibo"))
        cliente = str(val(r, "nro_cliente") or "").strip()
        facturas = [f.strip() for f in re.split(r"[/,;]", str(crudo_factura or ""))
                    if f.strip()]

        if not facturas:
            # Sin factura no es una cobranza procesable. Si además no tiene
            # nada, es fila de relleno (se ignora en silencio). Si tenía
            # plata o cliente, es una fila sospechosa → se reporta.
            if efectivo or cheque or total or cliente:
                descartadas.append({
                    "fila": r,
                    "cliente": cliente,
                    "cobrado": total or (efectivo + cheque),
                    "motivo": "Fila con datos pero sin Nº de factura",
                })
            continue
        fecha_raw = val(r, "fecha")
        fecha = fecha_raw if isinstance(fecha_raw, datetime) else None

        filas.append(FilaCobranza(
            fila_excel=r,
            fecha=fecha,
            nro_recibo=str(val(r, "nro_recibo") or "").strip(),
            nro_cliente=cliente,
            facturas=facturas,
            descuento_flag=_parse_flag_descuento(val(r, "descuento")),
            nro_cheque=str(val(r, "nro_cheque") or "").strip(),
            observaciones=str(val(r, "observaciones") or "").strip(),
            efectivo=efectivo,
            cheque=cheque,
            total_recibo=total,
        ))
    return filas, descartadas


# =====================================================================
# Índice de facturas desde Contabilium (capa de red)
# =====================================================================

def construir_indice_facturas(
    session: api_loader.ApiSession,
    fecha_desde: str,
    fecha_hasta: str,
) -> tuple[api_loader.ApiSession, dict]:
    """Pagina /api/comprobantes/search y arma un índice por Nº de factura.

    Devuelve (sesión, indice) donde indice tiene forma:
        {
          "fuerte": {norm_factura: [comprobante, ...]},
          "sufijo": {sufijo_num: [comprobante, ...]},
        }
    `comprobante` = {id, numero, total, id_cliente, razon_social, tipo}.

    OJO: `fuerte` mapea a una LISTA, no a un solo comprobante. En Contabilium
    las facturas (FAC) y las notas de crédito (NCF/NCT/NCE) numeran por
    secuencias SEPARADAS, así que un mismo `Numero` (ej. "A-00033352") puede
    existir como factura Y como NC. Guardar ambos y elegir en `_buscar_factura`
    (que prefiere la factura) evita que una NC reciente pise a la factura vieja
    del mismo número — bug reportado por Valeria 2026-07-07.
    El `total` sale de `ImporteTotalNeto`, que es el total CON IVA — lo que
    efectivamente se cobra. OJO: Contabilium tiene los nombres invertidos
    respecto de lo intuitivo: `ImporteTotalNeto` = con IVA, y
    `ImporteTotalBruto` = sin IVA (validado: Neto = Bruto × 1,22 en FAC y
    NCF). El `Saldo` (también con IVA) NO viene en el search: se pide en
    detalle con `obtener_saldo_factura` solo para las facturas matcheadas.

    Caveat conocido (ver facturador.cargar_facturas_via_api): el server
    ignora ?refExterna y no filtra por número, hay que paginar el rango y
    filtrar client-side. Usar un rango acotado (ej. últimos 60-90 días).
    """
    path = f"/api/comprobantes/search?fechaDesde={fecha_desde}&fechaHasta={fecha_hasta}"
    session, items = api_loader.api_paginate(session, path)

    fuerte: dict[str, list[dict]] = {}
    sufijo: dict[str, list[dict]] = {}
    for it in items:
        numero = str(it.get("Numero") or "").strip()
        if not numero:
            continue
        comp = {
            "id": it.get("Id") or it.get("ID") or 0,
            "numero": numero,
            "total": api_loader.parse_monto_uy(it.get("ImporteTotalNeto")),
            "id_cliente": it.get("IdCliente"),
            "razon_social": str(it.get("RazonSocial") or "").strip(),
            "tipo": str(it.get("TipoFc") or "").strip(),
        }
        # Acumular (no pisar): puede haber FAC y NC con el mismo número.
        fuerte.setdefault(_norm_factura(numero), []).append(comp)
        suf = _sufijo_num(numero)
        if suf:
            sufijo.setdefault(suf, []).append(comp)
    return session, {"fuerte": fuerte, "sufijo": sufijo}


def _es_factura(comp: dict) -> bool:
    """True si el comprobante es una factura (no una nota de crédito)."""
    return str(comp.get("tipo") or "").upper() not in TIPOS_NOTA_CREDITO


def obtener_saldo_factura(
    session: api_loader.ApiSession, id_comprobante
) -> tuple[api_loader.ApiSession, float]:
    """GET /api/comprobantes/?id= → Saldo pendiente (bruto). 0 = cobrada."""
    session, payload = api_loader.api_get(
        session, f"/api/comprobantes/?id={id_comprobante}"
    )
    saldo = api_loader.parse_monto_uy(payload.get("Saldo")) if isinstance(payload, dict) else 0.0
    return session, saldo


def _buscar_factura(numero: str, indice: dict) -> tuple[dict | None, str]:
    """Resuelve un Nº de factura contra el índice.

    Devuelve (comprobante, nota). `comprobante` None si no matchea.
    `nota` describe cómo matcheó (para trazabilidad en el reporte).

    Ante una colisión de número (una FAC y una NC con el mismo "A-000..."),
    SIEMPRE prefiere la factura: el vendedor cobra la factura, no la NC. Si el
    número solo matchea NC(s), devuelve una para que `analizar_fila` la marque
    REVISAR con el mensaje 'es una Nota de Crédito' (el vendedor se equivocó de
    número). Los números de factura son únicos dentro de su serie, así que
    preferir FAC desambigua sin necesitar el cliente.
    """
    fuerte = indice.get("fuerte", {})
    candidatos = fuerte.get(_norm_factura(numero), [])
    if candidatos:
        facturas = [c for c in candidatos if _es_factura(c)]
        if len(facturas) == 1:
            return facturas[0], ""
        if len(facturas) > 1:
            return None, "ambiguo: varias facturas con ese número"
        return candidatos[0], ""  # solo NC(s) → se marcará REVISAR

    # Respaldo por sufijo numérico, solo si es inequívoco (prefiriendo FAC).
    candidatos = indice.get("sufijo", {}).get(_sufijo_num(numero), [])
    facturas = [c for c in candidatos if _es_factura(c)]
    if len(facturas) == 1:
        return facturas[0], "match por número (serie difiere de formato)"
    if len(facturas) > 1:
        return None, "ambiguo: varios comprobantes con ese número"
    if len(candidatos) == 1:
        return candidatos[0], "match por número (serie difiere de formato)"
    if len(candidatos) > 1:
        return None, "ambiguo: varios comprobantes con ese número"
    return None, "factura no encontrada en el rango consultado"


# =====================================================================
# Análisis (lógica pura)
# =====================================================================

def analizar_fila(
    fila: FilaCobranza,
    indice: dict,
    tolerancia: float = TOLERANCIA_DEFAULT,
) -> ResultadoFila:
    """Analiza una fila y devuelve su ResultadoFila (OK / REVISAR + motivos).

    NO consulta saldos (eso lo agrega el orquestador para las facturas
    encontradas). Trabaja sobre `indice` en memoria: función pura.
    """
    res = ResultadoFila(fila=fila, estado=ESTADO_OK)

    # --- Resolver factura(s) ---
    encontrados: list[dict] = []
    for num in fila.facturas:
        comp, nota = _buscar_factura(num, indice)
        res.facturas_detalle.append({
            "numero_planilla": num,
            "comprobante": comp,
            "nota": nota,
            "saldo": None,
        })
        if comp is None:
            res.estado = ESTADO_REVISAR
            res.motivos.append(f"{num}: {nota}")
        else:
            # Detectar que el "número de factura" es en realidad una Nota de
            # Crédito (el vendedor se equivocó): no se cobra contra una NC.
            if str(comp.get("tipo") or "").upper() in TIPOS_NOTA_CREDITO:
                res.estado = ESTADO_REVISAR
                res.motivos.append(
                    f"{num}: es una Nota de Crédito ({comp['tipo']}), no una "
                    "factura. No se puede cobrar contra una NC — verificar el "
                    "número con el vendedor."
                )
                # No lo sumamos a `encontrados`: su total es negativo y
                # distorsionaría el cálculo de NC/cobro esperado.
                continue
            encontrados.append(comp)
            if nota:
                res.motivos.append(f"{num}: {nota}")

    if not encontrados:
        return res  # nada más que calcular

    res.total_factura = sum(c["total"] for c in encontrados)

    # --- Consistencia interna de la planilla ---
    if fila.total_recibo and abs(fila.total_recibo - (fila.efectivo + fila.cheque)) > 0.5:
        res.estado = ESTADO_REVISAR
        res.motivos.append(
            f"Total Recibo ({fila.total_recibo:,.0f}) ≠ efectivo+cheque "
            f"({fila.efectivo + fila.cheque:,.0f})"
        )
    if fila.cheque and not fila.nro_cheque:
        res.estado = ESTADO_REVISAR
        res.motivos.append(
            "Cobro con cheque sin Nº de cheque: no se puede imputar el "
            "cheque ya cargado en Contabilium sin su número"
        )

    # --- Multi-factura: se reporta pero requiere distribución manual ---
    if fila.es_multifactura:
        res.estado = ESTADO_REVISAR
        res.motivos.append(
            "Multi-factura: la planilla no indica cuánto se cobra a cada "
            "factura; requiere distribución manual"
        )
        # Igual dejamos calculado el descuento agregado como referencia.

    # --- Detectar entrega / pago parcial ---
    # Una "entrega" es un pago parcial de lo que el cliente adeuda: NO lleva
    # NC del 10% y no tiene un "cobro esperado" de 90/100%. Se detecta por la
    # palabra "entrega" en Observaciones, o porque (sin flag explícito) el
    # cobro no se acerca ni al 90% ni al 100% del total. Sin este chequeo, la
    # inferencia por cercanía asumía 10% para cualquier monto < 100% (bug:
    # una entrega de $5.000 caía como "10% asumido").
    cobrado = fila.cobrado
    esp_90 = round(res.total_factura * (1 - TASA_DESCUENTO), 2)
    esp_100 = round(res.total_factura, 2)
    obs_entrega = "entrega" in _norm(fila.observaciones)
    cerca_90 = abs(cobrado - esp_90) <= tolerancia
    cerca_100 = abs(cobrado - esp_100) <= tolerancia

    if not fila.es_multifactura and (
        obs_entrega or (fila.descuento_flag is None and not cerca_90 and not cerca_100)
    ):
        res.es_pago_parcial = True
        res.estado = ESTADO_REVISAR
        etiqueta = (
            "Entrega / pago parcial" if obs_entrega
            else "El cobro no coincide con el 90% ni el 100% de la factura"
        )
        res.motivos.append(
            f"{etiqueta}: cobrado {cobrado:,.0f} de una factura de "
            f"{esp_100:,.0f} (con IVA). No se asume descuento; definir "
            "manualmente si corresponde NC y cómo imputar."
        )
        return res

    # --- Determinar si aplica el 10% ---
    if fila.descuento_flag is not None:
        res.descuento_aplica = fila.descuento_flag
    elif cerca_90:
        res.descuento_aplica = True
        res.descuento_asumido = True
        res.motivos.append(
            "Descuento no especificado; asumido 10% (el cobro coincide con el 90%)"
        )
    elif cerca_100:
        res.descuento_aplica = False
        res.descuento_asumido = True
        res.motivos.append(
            "Descuento no especificado; asumido pago total (el cobro coincide con el 100%)"
        )
    else:
        # Multi-factura sin flag: no se puede inferir por monto (el cobro se
        # reparte). Ya está en REVISAR por multi-factura.
        res.descuento_aplica = False
        res.descuento_asumido = True
        res.motivos.append("Descuento no determinado por monto (multi-factura)")

    # --- Calcular esperado ---
    if res.descuento_aplica:
        res.nc_esperada = round(res.total_factura * TASA_DESCUENTO, 2)
        res.cobro_esperado = round(res.total_factura - res.nc_esperada, 2)
    else:
        res.nc_esperada = 0.0
        res.cobro_esperado = round(res.total_factura, 2)

    # --- Comparar cobro real vs esperado (solo si es 1 factura clara) ---
    res.diferencia = round(fila.cobrado - res.cobro_esperado, 2)
    if not fila.es_multifactura:
        if abs(res.diferencia) > tolerancia:
            res.estado = ESTADO_REVISAR
            res.motivos.append(
                f"Cobrado {fila.cobrado:,.0f} vs esperado {res.cobro_esperado:,.0f} "
                f"(dif {res.diferencia:+,.0f}, tolerancia ±{tolerancia:,.0f})"
            )

    return res


def analizar(
    filas: list[FilaCobranza],
    indice: dict,
    tolerancia: float = TOLERANCIA_DEFAULT,
) -> list[ResultadoFila]:
    """Analiza todas las filas (lógica pura, sin red)."""
    return [analizar_fila(f, indice, tolerancia) for f in filas]


def verificar_saldos(
    session: api_loader.ApiSession,
    resultados: list[ResultadoFila],
) -> api_loader.ApiSession:
    """Consulta el saldo pendiente de cada factura encontrada y lo anota.

    Muta `resultados` in-place (rellena `saldo` en facturas_detalle y
    `res.saldo_factura`) y agrega motivos:
      - saldo 0 → factura YA cobrada (riesgo de doble cobro) → REVISAR.
      - 0 < saldo < total → pago parcial previo (entrega) → REVISAR.

    Se hace en un paso separado (con red) para no ensuciar `analizar_fila`,
    que es pura. Solo pega a la API por las facturas efectivamente
    encontradas, cacheando por id para no repetir en multi-factura.
    """
    cache: dict = {}
    for res in resultados:
        saldos_ok = True
        for det in res.facturas_detalle:
            comp = det.get("comprobante")
            if not comp:
                continue
            cid = comp["id"]
            if cid not in cache:
                session, cache[cid] = obtener_saldo_factura(session, cid)
            saldo = cache[cid]
            det["saldo"] = saldo
            if saldo <= 0.5:
                res.estado = ESTADO_REVISAR
                res.motivos.append(
                    f"{det['numero_planilla']}: factura ya cobrada "
                    f"(saldo 0) — riesgo de doble cobro"
                )
                saldos_ok = False
            elif saldo + 0.5 < comp["total"]:
                res.motivos.append(
                    f"{det['numero_planilla']}: pago parcial previo "
                    f"(saldo {saldo:,.0f} de {comp['total']:,.0f})"
                )
        if saldos_ok:
            res.saldo_factura = sum(
                (d.get("saldo") or 0.0) for d in res.facturas_detalle
            )
    return session


# =====================================================================
# Reporte
# =====================================================================

def resultados_a_dataframe(resultados: list[ResultadoFila]) -> pd.DataFrame:
    """Aplana los resultados a un DataFrame listo para mostrar/descargar."""
    filas = []
    for res in resultados:
        f = res.fila
        if res.es_pago_parcial:
            desc_label = "Entrega/parcial"
        else:
            desc_label = ("10%" if res.descuento_aplica else "No") + (
                " (asumido)" if res.descuento_asumido else ""
            )
        filas.append({
            "Fila": f.fila_excel,
            "Fecha": f.fecha.date() if f.fecha else None,
            "Cliente": f.nro_cliente,
            "Factura(s)": " / ".join(f.facturas),
            "Total Factura": round(res.total_factura, 2),
            "Descuento": desc_label,
            "NC Esperada": None if res.es_pago_parcial else res.nc_esperada,
            "Cobro Esperado": None if res.es_pago_parcial else res.cobro_esperado,
            "Efectivo": f.efectivo,
            "Cheque": f.cheque,
            "Nº Cheque": f.nro_cheque,
            "Cobrado": f.cobrado,
            "Diferencia": res.diferencia,
            "Estado": res.estado,
            "Motivo": " | ".join(res.motivos),
        })
    return pd.DataFrame(filas)
