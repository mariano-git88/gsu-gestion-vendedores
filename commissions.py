"""
Lógica pura de cálculo de comisiones GSU.

No depende de Streamlit. Importable desde cualquier script o notebook.
Las reglas implementadas acá replican exactamente las definidas en
CLAUDE.md y en _learning/decisions.md (sesión 2026-04-09).

Reglas:
  - Comisión por venta = 2,35% × (Total ÷ 1,22)  [neto de IVA 22%]
  - Comisión por cobranza = 3% × Importe Total Neto
  - Redondeo: hacia arriba al peso (math.ceil), sin decimales
  - Vendedores excluidos siempre: OPJESICA, OPVALERIA
  - Estado excluido: Cancelada
  - Vendedor sin clientes vinculados → se excluye completo (ni venta ni cobranza)
  - Cobranza con código inexistente en clientes.xlsx → se asigna a MARIO
  - Cobranza de cliente existente sin Vendedor Asignado → se descarta
  - Toda operación se asume en UYU
"""

import io
import math
from collections import defaultdict
from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ----------------------------- CONSTANTES -----------------------------

TASA_VENTA = 0.0235
TASA_COBRANZA = 0.03
DIVISOR_IVA = 1.22

VENDEDORES_EXCLUIDOS_OP = {
    "OPJESICA@SUPRABOND.COM.UY",
    "OPVALERIA@SUPRABOND.COM.UY",
}
ESTADOS_EXCLUIDOS = {"Cancelada"}
VENDEDOR_HUERFANAS = "MARIO@SUPRABOND.COM.UY"


# ----------------------------- HELPERS -----------------------------

def parse_eu_number(s):
    """Parsea string en formato europeo ('1.537,20') a float."""
    if isinstance(s, (int, float)):
        return float(s)
    if s is None or s == "":
        return 0.0
    return float(str(s).replace(".", "").replace(",", "."))


def _require_columns(headers_set, required, sheet_name):
    missing = required - headers_set
    if missing:
        raise ValueError(
            f"En la hoja '{sheet_name}' faltan columnas requeridas: {sorted(missing)}"
        )


# ----------------------------- LOADERS -----------------------------

def load_clientes(file_or_path):
    """
    Devuelve (mapa_codigo_a_vendedor, set_de_vendedores_validos).
    'file_or_path' puede ser una ruta o un objeto file-like (UploadedFile de Streamlit).
    """
    wb = openpyxl.load_workbook(file_or_path, data_only=True)
    if "Clientes" not in wb.sheetnames:
        raise ValueError("clientes.xlsx debe tener una hoja llamada 'Clientes'")
    ws = wb["Clientes"]
    headers = {c.value: i for i, c in enumerate(ws[1])}
    _require_columns(set(headers), {"Codigo", "Vendedor Asignado"}, "Clientes")

    mapa = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = row[headers["Codigo"]]
        if cod is None:
            continue
        mapa[cod] = row[headers["Vendedor Asignado"]]

    valid_vendors = {v for v in mapa.values() if v is not None}
    return mapa, valid_vendors


def load_ventas(file_or_path, valid_vendors):
    """Lee ventas.xlsx y devuelve un dict con brutas/netas/detalle/excluidas por vendedor."""
    wb = openpyxl.load_workbook(file_or_path, data_only=True)
    if "OrdenesVenta" not in wb.sheetnames:
        raise ValueError("ventas.xlsx debe tener una hoja llamada 'OrdenesVenta'")
    ws = wb["OrdenesVenta"]
    headers = {c.value: i for i, c in enumerate(ws[1])}
    _require_columns(
        set(headers),
        {"Vendedor", "Estado", "Total", "Numero Orden", "Fecha Creacion", "Comprador", "Moneda"},
        "OrdenesVenta",
    )

    ventas_brutas = defaultdict(float)
    ventas_netas = defaultdict(float)
    detalle = defaultdict(list)
    excluidas = {"vendedor_op": 0, "cancelada": 0, "vendedor_invalido": 0}
    bruto_excluido_invalido = defaultdict(float)
    monedas_no_uyu = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        vend = row[headers["Vendedor"]]
        estado = row[headers["Estado"]]
        moneda = row[headers["Moneda"]]

        # Validación de moneda — decisión 2026-04-09: todo es UYU
        if moneda is not None and str(moneda).upper() != "UYU":
            monedas_no_uyu.append((row[headers["Numero Orden"]], moneda))
            # No se procesa, se reporta en alertas
            continue

        if vend in VENDEDORES_EXCLUIDOS_OP:
            excluidas["vendedor_op"] += 1
            continue
        if estado in ESTADOS_EXCLUIDOS:
            excluidas["cancelada"] += 1
            continue
        if vend not in valid_vendors:
            excluidas["vendedor_invalido"] += 1
            bruto_excluido_invalido[vend] += parse_eu_number(row[headers["Total"]])
            continue

        total_b = parse_eu_number(row[headers["Total"]])
        total_n = total_b / DIVISOR_IVA
        ventas_brutas[vend] += total_b
        ventas_netas[vend] += total_n
        detalle[vend].append({
            "numero": row[headers["Numero Orden"]],
            "fecha": str(row[headers["Fecha Creacion"]]) if row[headers["Fecha Creacion"]] else "",
            "comprador": row[headers["Comprador"]],
            "estado": estado,
            "total_bruto": total_b,
            "total_neto": total_n,
        })

    return {
        "brutas": dict(ventas_brutas),
        "netas": dict(ventas_netas),
        "detalle": dict(detalle),
        "excluidas": excluidas,
        "bruto_excluido_invalido": dict(bruto_excluido_invalido),
        "monedas_no_uyu": monedas_no_uyu,
    }


def load_cobranzas(file_or_path, mapa_clientes):
    """Lee cobranzas.xlsx y agrupa por vendedor, aplicando las reglas de huérfanas y descartes."""
    wb = openpyxl.load_workbook(file_or_path, data_only=True)
    if "Cobranzas" not in wb.sheetnames:
        raise ValueError("cobranzas.xlsx debe tener una hoja llamada 'Cobranzas'")
    ws = wb["Cobranzas"]
    headers = {c.value: i for i, c in enumerate(ws[1])}
    _require_columns(
        set(headers),
        {"Codigo", "Importe Total Neto", "Razon Social", "Numero", "Fecha", "Moneda"},
        "Cobranzas",
    )

    por_vend = defaultdict(float)
    detalle = defaultdict(list)
    huerfanas_a_mario = []
    descartadas_sin_vendedor = []
    monedas_no_uyu = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        cod = row[headers["Codigo"]]
        imp = float(row[headers["Importe Total Neto"]] or 0)
        razon = row[headers["Razon Social"]]
        nro = row[headers["Numero"]]
        fecha = row[headers["Fecha"]]
        moneda = row[headers["Moneda"]]

        if moneda is not None and str(moneda).upper() != "UYU":
            monedas_no_uyu.append((nro, moneda))
            continue

        if cod not in mapa_clientes:
            por_vend[VENDEDOR_HUERFANAS] += imp
            detalle[VENDEDOR_HUERFANAS].append({
                "codigo": cod, "razon": razon, "numero": nro,
                "fecha": str(fecha) if fecha else "", "importe": imp,
                "asignacion": "huerfana_a_mario",
            })
            huerfanas_a_mario.append((cod, razon, nro, imp))
        elif mapa_clientes[cod] is None:
            descartadas_sin_vendedor.append((cod, razon, nro, imp))
        else:
            v = mapa_clientes[cod]
            por_vend[v] += imp
            detalle[v].append({
                "codigo": cod, "razon": razon, "numero": nro,
                "fecha": str(fecha) if fecha else "", "importe": imp,
                "asignacion": "directa",
            })

    return {
        "por_vend": dict(por_vend),
        "detalle": dict(detalle),
        "huerfanas_a_mario": huerfanas_a_mario,
        "descartadas_sin_vendedor": descartadas_sin_vendedor,
        "monedas_no_uyu": monedas_no_uyu,
    }


# ----------------------------- AJUSTE RETROACTIVO -----------------------------

# Tolerancia para detectar cambios reales de importe (evita falsos positivos
# por redondeos al exportar de Contabilium).
TOLERANCIA_IMPORTE = 0.005


def _index_cobranzas_por_numero(ws):
    """
    Indexa una hoja de cobranzas por el campo 'Numero'.
    Devuelve {numero: {numero, codigo, razon, importe, fecha}}.
    Las cobranzas con moneda distinta de UYU se ignoran (no deberían existir
    según la regla, pero si aparecen no rompen este flujo).
    """
    headers = {c.value: i for i, c in enumerate(ws[1])}
    _require_columns(
        set(headers),
        {"Codigo", "Importe Total Neto", "Razon Social", "Numero", "Fecha", "Moneda"},
        ws.title,
    )
    idx = {}
    duplicados = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        nro = row[headers["Numero"]]
        if nro is None:
            continue
        moneda = row[headers["Moneda"]]
        if moneda is not None and str(moneda).upper() != "UYU":
            continue
        item = {
            "numero": nro,
            "codigo": row[headers["Codigo"]],
            "razon": row[headers["Razon Social"]],
            "importe": float(row[headers["Importe Total Neto"]] or 0),
            "fecha": str(row[headers["Fecha"]]) if row[headers["Fecha"]] else "",
        }
        if nro in idx:
            duplicados.append(nro)
        idx[nro] = item
    return idx, duplicados


def compute_retroactive_adjustment(file_orig, file_updated, mapa_clientes):
    """
    Compara dos versiones de la planilla cobranzas.xlsx del MES ANTERIOR y
    calcula el ajuste retroactivo de comisiones a aplicar este mes.

    Args:
        file_orig: planilla del mes anterior tal como se usó la corrida pasada
        file_updated: planilla del mes anterior re-exportada hoy de Contabilium
        mapa_clientes: dict {codigo: vendedor} del MES CORRIENTE (clientes.xlsx
                       que cargó el usuario para la corrida actual)

    Devuelve un dict con:
        - cambios: lista de cambios detectados (nuevas, eliminadas, modificadas)
        - delta_importe_por_vendedor: dict vendedor -> delta total de importe
        - ajuste_comision_por_vendedor: dict vendedor -> delta_importe × 3%
        - vendedores_con_ajuste_negativo: dict vendedor -> ajuste negativo
        - huerfanas_a_mario: cambios cuyo código no está en clientes.xlsx
        - descartadas_sin_vendedor: cambios cuyo cliente no tiene Vendedor Asignado
        - duplicados_orig / duplicados_updated: número de cobranzas duplicadas
          (deberían ser 0; si > 0 es un dato sucio que vale la pena reportar)

    Identidad de cobranzas: campo 'Numero' (único en Contabilium — confirmado
    por Mariano, sesión 2026-04-09).

    Reglas aplicadas: ver _learning/decisions.md, entrada 2026-04-09 sobre
    ajuste retroactivo.
    """
    wb_o = openpyxl.load_workbook(file_orig, data_only=True)
    wb_u = openpyxl.load_workbook(file_updated, data_only=True)
    if "Cobranzas" not in wb_o.sheetnames:
        raise ValueError("La planilla 'original' debe tener una hoja 'Cobranzas'")
    if "Cobranzas" not in wb_u.sheetnames:
        raise ValueError("La planilla 'actualizada' debe tener una hoja 'Cobranzas'")

    orig_idx, dup_o = _index_cobranzas_por_numero(wb_o["Cobranzas"])
    updated_idx, dup_u = _index_cobranzas_por_numero(wb_u["Cobranzas"])

    cambios = []
    todos_numeros = set(orig_idx) | set(updated_idx)
    for nro in sorted(todos_numeros, key=lambda x: str(x)):
        o = orig_idx.get(nro)
        u = updated_idx.get(nro)

        if o is None and u is not None:
            cambios.append({
                "tipo": "nueva",
                "numero": nro,
                "codigo": u["codigo"],
                "razon": u["razon"],
                "importe_original": 0.0,
                "importe_nuevo": u["importe"],
                "delta_importe": u["importe"],
                "fecha": u["fecha"],
            })
        elif u is None and o is not None:
            cambios.append({
                "tipo": "eliminada",
                "numero": nro,
                "codigo": o["codigo"],
                "razon": o["razon"],
                "importe_original": o["importe"],
                "importe_nuevo": 0.0,
                "delta_importe": -o["importe"],
                "fecha": o["fecha"],
            })
        else:
            if abs(o["importe"] - u["importe"]) > TOLERANCIA_IMPORTE:
                cambios.append({
                    "tipo": "modificada",
                    "numero": nro,
                    "codigo": u["codigo"],  # versión más reciente
                    "razon": u["razon"],
                    "importe_original": o["importe"],
                    "importe_nuevo": u["importe"],
                    "delta_importe": u["importe"] - o["importe"],
                    "fecha": u["fecha"],
                })

    # Asignar cada cambio a un vendedor con las MISMAS reglas del mes corriente
    delta_importe_vend = defaultdict(float)
    huerfanas_a_mario = []
    descartadas = []

    for c in cambios:
        cod = c["codigo"]
        if cod not in mapa_clientes:
            delta_importe_vend[VENDEDOR_HUERFANAS] += c["delta_importe"]
            c["asignacion"] = f"huerfana_a_{VENDEDOR_HUERFANAS}"
            huerfanas_a_mario.append(c)
        elif mapa_clientes[cod] is None:
            c["asignacion"] = "descartada_sin_vendedor"
            descartadas.append(c)
        else:
            v = mapa_clientes[cod]
            delta_importe_vend[v] += c["delta_importe"]
            c["asignacion"] = v

    # Comisión = delta_importe × 3%
    ajuste_comision_vend = {
        v: importe * TASA_COBRANZA for v, importe in delta_importe_vend.items()
    }
    vendedores_con_ajuste_negativo = {
        v: monto for v, monto in ajuste_comision_vend.items() if monto < 0
    }

    return {
        "cambios": cambios,
        "delta_importe_por_vendedor": dict(delta_importe_vend),
        "ajuste_comision_por_vendedor": ajuste_comision_vend,
        "vendedores_con_ajuste_negativo": vendedores_con_ajuste_negativo,
        "huerfanas_a_mario": huerfanas_a_mario,
        "descartadas_sin_vendedor": descartadas,
        "duplicados_original": dup_o,
        "duplicados_actualizada": dup_u,
        "total_orig": sum(x["importe"] for x in orig_idx.values()),
        "total_actualizada": sum(x["importe"] for x in updated_idx.values()),
    }


def merge_commissions_with_adjustment(resumen_normal, ajuste):
    """
    Fusiona el resumen normal del mes con el ajuste retroactivo.

    Reglas (definidas en _learning/decisions.md, entrada 2026-04-09):
    - Si el ajuste para un vendedor es ≥ 0: se SUMA a su comisión bruta.
    - Si el ajuste es < 0: NO se descuenta. Queda registrado como alerta.
    - El total final se redondea hacia arriba al peso (igual que sin ajuste).
    - Si un vendedor solo aparece en el ajuste (no tiene operaciones este mes),
      igual se incluye con sus campos del mes corriente en cero.

    Devuelve una lista nueva de filas con campos extra:
        - ajuste_comision: el ajuste tal como salió del cálculo (puede ser negativo)
        - ajuste_aplicado: lo que efectivamente se sumó (>= 0; igual a ajuste si > 0, sino 0)
        - comision_bruta_con_ajuste: comision_bruta + ajuste_aplicado
        - comision_neta_con_ajuste: ceil(comision_bruta_con_ajuste)
    """
    ajuste_dict = ajuste["ajuste_comision_por_vendedor"]
    resumen_lookup = {r["vendedor"]: r for r in resumen_normal}

    todos = sorted(set(resumen_lookup) | set(ajuste_dict))
    nueva = []
    for v in todos:
        base = resumen_lookup.get(v) or {
            "vendedor": v,
            "ventas_brutas": 0.0,
            "ventas_netas": 0.0,
            "cobranzas": 0.0,
            "comision_venta": 0.0,
            "comision_cobranza": 0.0,
            "comision_bruta": 0.0,
            "comision_neta": 0,
        }

        ajuste_v = ajuste_dict.get(v, 0.0)
        ajuste_aplicado = ajuste_v if ajuste_v > 0 else 0.0

        bruta_total = base["comision_bruta"] + ajuste_aplicado
        neta_final = math.ceil(bruta_total)

        nueva.append({
            **base,
            "ajuste_comision": ajuste_v,
            "ajuste_aplicado": ajuste_aplicado,
            "comision_bruta_con_ajuste": bruta_total,
            "comision_neta_con_ajuste": neta_final,
        })
    return nueva


# ----------------------------- CÁLCULO -----------------------------

def compute_commissions(ventas, cobranzas):
    """Combina ventas y cobranzas y devuelve la lista de filas de la liquidación."""
    todos = sorted(set(ventas["netas"].keys()) | set(cobranzas["por_vend"].keys()))
    resumen = []
    for v in todos:
        vn = ventas["netas"].get(v, 0.0)
        vb = ventas["brutas"].get(v, 0.0)
        co = cobranzas["por_vend"].get(v, 0.0)
        com_v = vn * TASA_VENTA
        com_c = co * TASA_COBRANZA
        com_bruta = com_v + com_c
        com_neta = math.ceil(com_bruta)  # redondeo hacia arriba al peso
        resumen.append({
            "vendedor": v,
            "ventas_brutas": vb,
            "ventas_netas": vn,
            "cobranzas": co,
            "comision_venta": com_v,
            "comision_cobranza": com_c,
            "comision_bruta": com_bruta,
            "comision_neta": com_neta,
        })
    return resumen


# ----------------------------- BUILDERS DE OUTPUT -----------------------------

def _styles():
    return {
        "bold": Font(bold=True),
        "header_fill": PatternFill("solid", fgColor="305496"),
        "header_font": Font(bold=True, color="FFFFFF"),
        "money_fmt": "#,##0.00",
        "money0": "#,##0",
        "center": Alignment(horizontal="center", vertical="center"),
        "border": Border(
            left=Side(border_style="thin", color="999999"),
            right=Side(border_style="thin", color="999999"),
            top=Side(border_style="thin", color="999999"),
            bottom=Side(border_style="thin", color="999999"),
        ),
    }


def build_xlsx_bytes(resumen, ventas, cobranzas, periodo_label, ajuste=None):
    """
    Genera el .xlsx de liquidación y lo devuelve como BytesIO.

    Si ajuste no es None, se asume que `resumen` viene de
    merge_commissions_with_adjustment() y trae los campos extra
    `ajuste_comision`, `ajuste_aplicado`, `comision_neta_con_ajuste`.
    En ese caso, se agregan dos columnas y una hoja extra "Ajuste retroactivo".
    """
    has_ajuste = ajuste is not None
    s = _styles()
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen"

    last_col_letter = "J" if has_ajuste else "H"
    ws["A1"] = "Liquidación de Comisiones — GSU"
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A2"] = f"Período: {periodo_label}"
    ws["A2"].font = s["bold"]
    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A3"] = f"Generado: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells(f"A3:{last_col_letter}3")

    headers = [
        "Vendedor",
        "Ventas brutas (c/IVA)",
        "Ventas netas (s/IVA)",
        "Cobranzas",
        "Comisión venta (2,35%)",
        "Comisión cobranza (3%)",
        "Comisión bruta",
        "Comisión neta a pagar",
    ]
    if has_ajuste:
        headers += [
            "Ajuste mes anterior",
            "TOTAL a pagar (con ajuste)",
        ]
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=5, column=j, value=h)
        c.font = s["header_font"]
        c.fill = s["header_fill"]
        c.alignment = s["center"]
        c.border = s["border"]

    row = 6
    for r in resumen:
        ws.cell(row=row, column=1, value=r["vendedor"]).border = s["border"]
        cells = [
            (2, r["ventas_brutas"], s["money_fmt"]),
            (3, r["ventas_netas"], s["money_fmt"]),
            (4, r["cobranzas"], s["money_fmt"]),
            (5, r["comision_venta"], s["money_fmt"]),
            (6, r["comision_cobranza"], s["money_fmt"]),
            (7, r["comision_bruta"], s["money_fmt"]),
            (8, r["comision_neta"], s["money0"]),
        ]
        if has_ajuste:
            cells += [
                (9, r.get("ajuste_aplicado", 0.0), s["money_fmt"]),
                (10, r.get("comision_neta_con_ajuste", r["comision_neta"]), s["money0"]),
            ]
        for col, val, fmt in cells:
            c = ws.cell(row=row, column=col, value=val)
            c.number_format = fmt
            c.border = s["border"]
        # Resaltar la columna de pago final
        pago_col = 10 if has_ajuste else 8
        ws.cell(row=row, column=pago_col).font = s["bold"]
        row += 1

    total_row = row
    ws.cell(row=total_row, column=1, value="TOTAL").font = s["bold"]
    ws.cell(row=total_row, column=1).border = s["border"]
    sum_cols = (2, 3, 4, 5, 6, 7)
    for col in sum_cols:
        letra = get_column_letter(col)
        c = ws.cell(row=total_row, column=col, value=f"=SUM({letra}6:{letra}{row-1})")
        c.number_format = s["money_fmt"]
        c.font = s["bold"]
        c.border = s["border"]
    c = ws.cell(row=total_row, column=8, value=f"=SUM(H6:H{row-1})")
    c.number_format = s["money0"]
    c.font = s["bold"]
    c.border = s["border"]
    if has_ajuste:
        c = ws.cell(row=total_row, column=9, value=f"=SUM(I6:I{row-1})")
        c.number_format = s["money_fmt"]
        c.font = s["bold"]
        c.border = s["border"]
        c = ws.cell(row=total_row, column=10, value=f"=SUM(J6:J{row-1})")
        c.number_format = s["money0"]
        c.font = Font(bold=True, size=12)
        c.border = s["border"]
    else:
        ws.cell(row=total_row, column=8).font = Font(bold=True, size=12)

    widths = [32, 22, 22, 18, 22, 24, 18, 22]
    if has_ajuste:
        widths += [22, 26]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ---- Notas ----
    ws_n = wb.create_sheet("Notas")
    ws_n["A1"] = "Notas de la corrida"
    ws_n["A1"].font = Font(bold=True, size=14)

    huerfanas = cobranzas["huerfanas_a_mario"]
    descartadas = cobranzas["descartadas_sin_vendedor"]
    bruto_inv = ventas["bruto_excluido_invalido"]

    notas = [
        "",
        "REGLAS DE CÁLCULO APLICADAS",
        f"  · Comisión por venta = {TASA_VENTA*100:.2f}% × (Total ÷ {DIVISOR_IVA})  [neto de IVA]",
        f"  · Comisión por cobranza = {TASA_COBRANZA*100:.2f}% × Importe Total Neto",
        "  · Comisión neta = REDONDEO HACIA ARRIBA al peso (sin decimales)",
        "",
        "FILTROS APLICADOS A VENTAS",
        f"  · {ventas['excluidas']['vendedor_op']} líneas excluidas por Vendedor = OPJESICA / OPVALERIA",
        f"  · {ventas['excluidas']['cancelada']} líneas excluidas por Estado = Cancelada",
        f"  · {ventas['excluidas']['vendedor_invalido']} líneas excluidas por Vendedor sin clientes vinculados",
    ]
    for v, m in bruto_inv.items():
        notas.append(f"      - {v}: bruto excluido ${m:,.2f}")
    notas += [
        "",
        "REASIGNACIONES Y DESCARTES EN COBRANZAS",
        f"  · {len(huerfanas)} cobranzas con código inexistente en clientes.xlsx → reasignadas a {VENDEDOR_HUERFANAS}:",
    ]
    for cod, razon, nro, imp in huerfanas:
        notas.append(f"      - {cod}  {razon}  Nº {nro}  ${imp:,.2f}")
    notas += [f"  · {len(descartadas)} cobranzas descartadas (cliente existe pero sin Vendedor Asignado):"]
    for cod, razon, nro, imp in descartadas:
        notas.append(f"      - {cod}  {razon}  Nº {nro}  ${imp:,.2f}")

    # Alertas de moneda no UYU (no debería pasar)
    if ventas.get("monedas_no_uyu") or cobranzas.get("monedas_no_uyu"):
        notas += ["", "ALERTA: se encontraron operaciones con moneda distinta de UYU"]
        for nro, mon in ventas.get("monedas_no_uyu", []):
            notas.append(f"  · Venta {nro}: moneda={mon}")
        for nro, mon in cobranzas.get("monedas_no_uyu", []):
            notas.append(f"  · Cobranza {nro}: moneda={mon}")

    if has_ajuste:
        notas += [
            "",
            "AJUSTE RETROACTIVO DEL MES ANTERIOR",
            f"  · {len(ajuste['cambios'])} cambios detectados al comparar las dos versiones de cobranzas del mes anterior.",
            f"  · Vendedores con ajuste positivo (sumado al pago): "
            + ", ".join(
                f"{v}=+${m:,.2f}"
                for v, m in ajuste["ajuste_comision_por_vendedor"].items()
                if m > 0
            )
            or "  · Vendedores con ajuste positivo: ninguno",
        ]
        if ajuste["vendedores_con_ajuste_negativo"]:
            notas += [
                f"  · ALERTA — Vendedores con ajuste NEGATIVO (NO descontado, requiere revisión manual):",
            ]
            for v, m in ajuste["vendedores_con_ajuste_negativo"].items():
                notas.append(f"      - {v}: ${m:,.2f}")

    for i, t in enumerate(notas, 2):
        ws_n.cell(row=i, column=1, value=t)
    ws_n.column_dimensions["A"].width = 110

    # ---- Hoja Ajuste retroactivo (solo si hay ajuste) ----
    if has_ajuste:
        ws_a = wb.create_sheet("Ajuste retroactivo")
        ws_a["A1"] = "Ajuste retroactivo del mes anterior"
        ws_a["A1"].font = Font(bold=True, size=14)
        ws_a.merge_cells("A1:H1")

        ws_a["A3"] = "Total cobranzas en versión ORIGINAL del mes anterior:"
        ws_a["F3"] = ajuste.get("total_orig", 0.0)
        ws_a["F3"].number_format = s["money_fmt"]
        ws_a["A4"] = "Total cobranzas en versión ACTUALIZADA del mes anterior:"
        ws_a["F4"] = ajuste.get("total_actualizada", 0.0)
        ws_a["F4"].number_format = s["money_fmt"]
        ws_a["A5"] = "Delta total (actualizada − original):"
        ws_a["F5"] = ajuste.get("total_actualizada", 0) - ajuste.get("total_orig", 0)
        ws_a["F5"].number_format = s["money_fmt"]
        ws_a["F5"].font = s["bold"]

        # Encabezado de detalle
        ws_a["A7"] = "Detalle de cambios"
        ws_a["A7"].font = Font(bold=True, size=12)
        h_a = ["Tipo", "Nº Cobranza", "Código", "Razón Social", "Importe original", "Importe nuevo", "Delta importe", "Asignación"]
        for j, h in enumerate(h_a, 1):
            c = ws_a.cell(row=8, column=j, value=h)
            c.font = s["header_font"]
            c.fill = s["header_fill"]
            c.border = s["border"]
        rr = 9
        for c_row in ajuste["cambios"]:
            ws_a.cell(row=rr, column=1, value=c_row["tipo"]).border = s["border"]
            ws_a.cell(row=rr, column=2, value=c_row["numero"]).border = s["border"]
            ws_a.cell(row=rr, column=3, value=c_row["codigo"]).border = s["border"]
            ws_a.cell(row=rr, column=4, value=c_row["razon"]).border = s["border"]
            for col, key in [(5, "importe_original"), (6, "importe_nuevo"), (7, "delta_importe")]:
                c = ws_a.cell(row=rr, column=col, value=c_row[key])
                c.number_format = s["money_fmt"]
                c.border = s["border"]
            ws_a.cell(row=rr, column=8, value=c_row["asignacion"]).border = s["border"]
            rr += 1

        # Resumen por vendedor
        rr += 2
        ws_a.cell(row=rr, column=1, value="Resumen del ajuste por vendedor").font = Font(bold=True, size=12)
        rr += 1
        h_r = ["Vendedor", "Delta importe", "Ajuste comisión (×3%)", "Estado"]
        for j, h in enumerate(h_r, 1):
            c = ws_a.cell(row=rr, column=j, value=h)
            c.font = s["header_font"]
            c.fill = s["header_fill"]
            c.border = s["border"]
        rr += 1
        for v in sorted(ajuste["ajuste_comision_por_vendedor"]):
            delta_imp = ajuste["delta_importe_por_vendedor"][v]
            ajuste_com = ajuste["ajuste_comision_por_vendedor"][v]
            estado = "APLICADO (positivo)" if ajuste_com > 0 else "NO APLICADO (negativo, alerta)"
            ws_a.cell(row=rr, column=1, value=v).border = s["border"]
            c = ws_a.cell(row=rr, column=2, value=delta_imp)
            c.number_format = s["money_fmt"]; c.border = s["border"]
            c = ws_a.cell(row=rr, column=3, value=ajuste_com)
            c.number_format = s["money_fmt"]; c.border = s["border"]
            ws_a.cell(row=rr, column=4, value=estado).border = s["border"]
            rr += 1

        for i, w in enumerate([14, 18, 14, 38, 18, 18, 18, 28], 1):
            ws_a.column_dimensions[get_column_letter(i)].width = w

    # ---- Hoja por vendedor ----
    for r in resumen:
        v = r["vendedor"]
        nombre_hoja = v.split("@")[0][:31]
        ws_v = wb.create_sheet(nombre_hoja)
        ws_v["A1"] = f"Detalle — {v}"
        ws_v["A1"].font = Font(bold=True, size=13)
        ws_v.merge_cells("A1:F1")
        ws_v["A2"] = f"Período: {periodo_label}"
        ws_v.merge_cells("A2:F2")

        pares = [
            ("Ventas brutas (c/IVA)", r["ventas_brutas"], s["money_fmt"]),
            ("Ventas netas (s/IVA)", r["ventas_netas"], s["money_fmt"]),
            ("Cobranzas", r["cobranzas"], s["money_fmt"]),
            ("Comisión venta (2,35%)", r["comision_venta"], s["money_fmt"]),
            ("Comisión cobranza (3%)", r["comision_cobranza"], s["money_fmt"]),
            ("Comisión bruta", r["comision_bruta"], s["money_fmt"]),
            ("Comisión neta a pagar", r["comision_neta"], s["money0"]),
        ]
        for k, (lbl, val, fmt) in enumerate(pares, start=4):
            ws_v.cell(row=k, column=1, value=lbl).font = s["bold"]
            c = ws_v.cell(row=k, column=2, value=val)
            c.number_format = fmt
        ws_v.cell(row=10, column=2).font = Font(bold=True, size=12)

        ws_v["A12"] = "Ventas del período"
        ws_v["A12"].font = Font(bold=True, size=11)
        h_v = ["Nº Orden", "Fecha", "Comprador", "Estado", "Total bruto", "Total neto"]
        for j, h in enumerate(h_v, 1):
            c = ws_v.cell(row=13, column=j, value=h)
            c.font = s["header_font"]
            c.fill = s["header_fill"]
            c.border = s["border"]
        rr = 14
        for d in ventas["detalle"].get(v, []):
            ws_v.cell(row=rr, column=1, value=d["numero"])
            ws_v.cell(row=rr, column=2, value=d["fecha"])
            ws_v.cell(row=rr, column=3, value=d["comprador"])
            ws_v.cell(row=rr, column=4, value=d["estado"])
            c = ws_v.cell(row=rr, column=5, value=d["total_bruto"])
            c.number_format = s["money_fmt"]
            c = ws_v.cell(row=rr, column=6, value=d["total_neto"])
            c.number_format = s["money_fmt"]
            rr += 1

        rr += 2
        ws_v.cell(row=rr, column=1, value="Cobranzas del período").font = Font(bold=True, size=11)
        rr += 1
        h_c = ["Código cliente", "Razón Social", "Nº Cobranza", "Fecha", "Importe", "Asignación"]
        for j, h in enumerate(h_c, 1):
            c = ws_v.cell(row=rr, column=j, value=h)
            c.font = s["header_font"]
            c.fill = s["header_fill"]
            c.border = s["border"]
        rr += 1
        for d in cobranzas["detalle"].get(v, []):
            ws_v.cell(row=rr, column=1, value=d["codigo"])
            ws_v.cell(row=rr, column=2, value=d["razon"])
            ws_v.cell(row=rr, column=3, value=d["numero"])
            ws_v.cell(row=rr, column=4, value=d["fecha"])
            c = ws_v.cell(row=rr, column=5, value=d["importe"])
            c.number_format = s["money_fmt"]
            ws_v.cell(row=rr, column=6, value=d["asignacion"])
            rr += 1

        for i, w in enumerate([18, 14, 42, 14, 18, 18], 1):
            ws_v.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def build_md(resumen, ventas, cobranzas, periodo_label, ajuste=None):
    """
    Genera el resumen narrativo en Markdown como string.

    Si ajuste no es None, se asume que `resumen` viene de
    merge_commissions_with_adjustment() y se incluyen las columnas y
    secciones del ajuste retroactivo.
    """
    has_ajuste = ajuste is not None
    pago_key = "comision_neta_con_ajuste" if has_ajuste else "comision_neta"
    total = sum(r[pago_key] for r in resumen)
    sorted_r = sorted(resumen, key=lambda r: -r[pago_key])
    top = sorted_r[0] if sorted_r else None
    bottom = sorted_r[-1] if sorted_r and len(sorted_r) > 1 else None

    huerfanas = cobranzas["huerfanas_a_mario"]
    descartadas = cobranzas["descartadas_sin_vendedor"]
    bruto_inv = ventas["bruto_excluido_invalido"]

    lines = [
        f"# Resumen — Liquidación de Comisiones GSU · {periodo_label}",
        "",
        f"**Generado:** {datetime.now().strftime('%Y-%m-%d %H:%M')}",
        f"**Período liquidado:** {periodo_label}",
        "",
        "---",
        "",
        "## Total a pagar",
        "",
        f"# **${total:,} UYU**",
        "",
        ("(Suma de comisiones netas del mes + ajustes retroactivos del mes anterior, redondeado hacia arriba al peso.)"
         if has_ajuste
         else "(Suma de comisiones netas redondeadas hacia arriba al peso.)"),
        "",
        "---",
        "",
        "## Liquidación por vendedor",
        "",
    ]
    if has_ajuste:
        lines += [
            "| Vendedor | Comisión mes (bruta) | Ajuste mes anterior | **TOTAL a pagar** |",
            "|---|---:|---:|---:|",
        ]
        for r in sorted_r:
            ajuste_str = f"+${r['ajuste_aplicado']:,.2f}" if r['ajuste_aplicado'] > 0 else (
                f"⚠ ${r['ajuste_comision']:,.2f} (no aplicado)" if r['ajuste_comision'] < 0 else "—"
            )
            lines.append(
                f"| {r['vendedor']} | {r['comision_bruta']:,.2f} | {ajuste_str} | "
                f"**${r['comision_neta_con_ajuste']:,}** |"
            )
        lines.append(f"| **TOTAL** | | | **${total:,}** |")
    else:
        lines += [
            "| Vendedor | Ventas netas | Cobranzas | Com. Venta (2,35%) | Com. Cobranza (3%) | **Comisión neta** |",
            "|---|---:|---:|---:|---:|---:|",
        ]
        for r in sorted_r:
            lines.append(
                f"| {r['vendedor']} | {r['ventas_netas']:,.2f} | {r['cobranzas']:,.2f} | "
                f"{r['comision_venta']:,.2f} | {r['comision_cobranza']:,.2f} | **${r['comision_neta']:,}** |"
            )
        lines.append(f"| **TOTAL** | | | | | **${total:,}** |")
    lines += [
        "",
        "---",
        "",
        "## Top y bottom",
        "",
    ]
    if top:
        lines.append(f"- **Top:** {top['vendedor']} — ${top[pago_key]:,}")
    if bottom:
        lines.append(f"- **Bottom:** {bottom['vendedor']} — ${bottom[pago_key]:,}")
    lines += [
        "",
        "---",
        "",
        "## Casos a revisar",
        "",
    ]
    if bruto_inv:
        lines.append("### Vendedores excluidos por no tener clientes vinculados")
        lines.append("")
        for v, m in bruto_inv.items():
            lines.append(f"- **{v}**: ${m:,.2f} brutos no contabilizados.")
        lines.append("")
    if huerfanas:
        lines.append(f"### Cobranzas con código inexistente — reasignadas a {VENDEDOR_HUERFANAS}")
        lines.append("")
        lines.append("| Código | Razón Social | Nº | Importe |")
        lines.append("|---|---|---|---:|")
        for cod, razon, nro, imp in huerfanas:
            lines.append(f"| {cod} | {razon} | {nro} | ${imp:,.2f} |")
        lines.append(f"| **Total** | | | **${sum(x[3] for x in huerfanas):,.2f}** |")
        lines.append("")
    if descartadas:
        lines.append("### Cobranzas descartadas (cliente existe sin Vendedor Asignado)")
        lines.append("")
        lines.append("| Código | Razón Social | Nº | Importe |")
        lines.append("|---|---|---|---:|")
        for cod, razon, nro, imp in descartadas:
            lines.append(f"| {cod} | {razon} | {nro} | ${imp:,.2f} |")
        lines.append("")
    if not (bruto_inv or huerfanas or descartadas):
        lines.append("Sin casos especiales en este período.")
        lines.append("")

    # ----- Sección del ajuste retroactivo -----
    if has_ajuste:
        lines += [
            "---",
            "",
            "## Ajuste retroactivo del mes anterior",
            "",
            f"Se compararon dos versiones de las cobranzas del mes anterior y se detectaron "
            f"**{len(ajuste['cambios'])} cambios**:",
            "",
            f"- Total cobranzas en versión **original**: ${ajuste.get('total_orig', 0):,.2f}",
            f"- Total cobranzas en versión **actualizada**: ${ajuste.get('total_actualizada', 0):,.2f}",
            f"- **Delta total** (actualizada − original): "
            f"${ajuste.get('total_actualizada', 0) - ajuste.get('total_orig', 0):+,.2f}",
            "",
        ]
        if ajuste["cambios"]:
            lines += [
                "### Detalle de cambios",
                "",
                "| Tipo | Nº | Código | Razón Social | Importe original | Importe nuevo | Delta | Asignación |",
                "|---|---|---|---|---:|---:|---:|---|",
            ]
            for c in ajuste["cambios"]:
                lines.append(
                    f"| {c['tipo']} | {c['numero']} | {c['codigo']} | {c['razon']} | "
                    f"${c['importe_original']:,.2f} | ${c['importe_nuevo']:,.2f} | "
                    f"${c['delta_importe']:+,.2f} | {c['asignacion']} |"
                )
            lines.append("")

        lines += [
            "### Ajuste por vendedor",
            "",
            "| Vendedor | Delta importe | Ajuste comisión (×3%) | Estado |",
            "|---|---:|---:|---|",
        ]
        for v in sorted(ajuste["ajuste_comision_por_vendedor"]):
            di = ajuste["delta_importe_por_vendedor"][v]
            ac = ajuste["ajuste_comision_por_vendedor"][v]
            estado = "Aplicado" if ac > 0 else "**NO aplicado** (alerta)"
            lines.append(f"| {v} | ${di:+,.2f} | ${ac:+,.2f} | {estado} |")
        lines.append("")

        if ajuste["vendedores_con_ajuste_negativo"]:
            lines += [
                "### ⚠ Alertas de ajuste negativo",
                "",
                "Estos vendedores tienen un ajuste retroactivo **negativo** (saldo a favor de "
                "Suprabond por anulaciones o reducciones de cobranzas pasadas). **No se les "
                "descontó automáticamente** — quedan con su comisión normal del mes. "
                "Revisar manualmente y decidir si compensar o no:",
                "",
            ]
            for v, m in ajuste["vendedores_con_ajuste_negativo"].items():
                lines.append(f"- **{v}**: ajuste = ${m:,.2f}")
            lines.append("")

    lines += [
        "---",
        "",
        "## Filtros aplicados",
        "",
        f"- {ventas['excluidas']['vendedor_op']} líneas excluidas por Vendedor ∈ {{OPJESICA, OPVALERIA}}",
        f"- {ventas['excluidas']['cancelada']} líneas excluidas por Estado = Cancelada",
        "- Todas las operaciones procesadas como UYU (regla fija — ver decisions.md 2026-04-09)",
    ]
    return "\n".join(lines)
