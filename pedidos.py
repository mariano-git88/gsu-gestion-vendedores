"""Lector de pedidos GSU (Fase 1).

Lee el adjunto `NOTA DE PEDIDO G.S.U.` que mandan los vendedores por mail
y devuelve los pedidos limpios y verificados, SIN tocar Contabilium.

El archivo es una plantilla fija: una hoja por pedido, el catalogo entero
precargado desde la fila 7, y el pedido real son solo las filas con
"Cantidad Pedida" > 0. La proteccion de hoja y las celdas combinadas se
ignoran solas al leer con openpyxl (no es encriptacion).

Control de totales validado contra archivos reales:
    suma(Sub-total) * 1.22  ==  TOTAL CON IVA   (IVA 22% UY)
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
import re

import openpyxl

IVA_UY = 0.22
# Tolerancia en pesos para dar por bueno el control de totales.
TOLERANCIA_TOTAL = 0.50

# Layout fijo de la plantilla (1-based, como openpyxl).
_CEL = {
    "cliente": (3, 3),       # C3
    "nro_cliente": (3, 5),   # E3
    "nro_vendedor": (3, 7),  # G3
    "nro_pedido": (2, 7),    # G2  (numero de pedido del vendedor)
    "cond_pago": (5, 3),     # C5  (texto libre, puede venir vacio)
    "total_con_iva": (5, 5), # E5
    "fecha": (4, 7),         # G4
}
_FILA_ITEMS = 7  # primera fila de datos (incluye encabezados de rubro)
# Columnas de la grilla de items.
_COL_CODIGO, _COL_DESC, _COL_UXC, _COL_PRECIO, _COL_CANT, _COL_SUBTOT = 2, 3, 4, 5, 6, 7


@dataclass
class Item:
    fila: int
    codigo: str
    descripcion: str
    und_x_caja: str
    precio_sin_iva: float
    cantidad: float
    subtotal: float

    @property
    def es_combo(self) -> bool:
        # Heuristica: los combos arrancan con "COM ". Se confirma en Fase 2
        # contra Contabilium; aca solo se marca para revision.
        return self.codigo.upper().startswith("COM ")


@dataclass
class Pedido:
    hoja: str
    cliente: str
    nro_cliente: str
    nro_vendedor: str
    nro_pedido: str
    cond_pago: str
    fecha: datetime | None
    total_con_iva_declarado: float | None
    items: list[Item] = field(default_factory=list)

    @property
    def suma_subtotales(self) -> float:
        return round(sum(i.subtotal for i in self.items), 2)

    @property
    def total_con_iva_calculado(self) -> float:
        return round(self.suma_subtotales * (1 + IVA_UY), 2)

    @property
    def total_ok(self) -> bool:
        """True si el total calculado cuadra con el declarado en el Excel."""
        if self.total_con_iva_declarado is None:
            return False
        return abs(self.total_con_iva_calculado - self.total_con_iva_declarado) <= TOLERANCIA_TOTAL

    @property
    def tiene_combos(self) -> bool:
        return any(i.es_combo for i in self.items)


def _txt(v) -> str:
    return "" if v is None else str(v).strip()


def _num(v) -> float | None:
    return float(v) if isinstance(v, (int, float)) else None


def codigo_cliente_candidatos(nro_cliente, cliente_texto) -> list[str]:
    """Códigos Contabilium candidatos ('0XXXX-C') para un pedido.

    Primero el Nro. Cliente (col E3). Como fallback, un número embebido
    al principio del nombre del cliente (col C3), porque muchas veces el
    vendedor escribe ahí el número: 'mucho texto' no matchea, pero
    '4016-barraca pirata' o '4016 barraca pirata' → 04016-C.

    Devuelve la lista de candidatos en orden de preferencia (sin
    duplicados). El llamador prueba cada uno contra el maestro y se
    queda con el primero que exista.
    """
    cands: list[str] = []

    def fmt(d) -> str | None:
        d = re.sub(r"\D", "", str(d or ""))
        return f"{int(d):05d}-C" if d else None

    c1 = fmt(nro_cliente)
    if c1:
        cands.append(c1)
    m = re.match(r"\s*0*(\d{2,6})\b", str(cliente_texto or ""))
    if m:
        c2 = fmt(m.group(1))
        if c2 and c2 not in cands:
            cands.append(c2)
    return cands


def leer_pedidos(file_or_path) -> list[Pedido]:
    """Lee el xlsx tal cual viene del mail y devuelve un Pedido por hoja.

    `file_or_path` puede ser una ruta o un file-like (ej. el upload de
    Streamlit). No modifica el archivo.
    """
    wb = openpyxl.load_workbook(file_or_path, data_only=True, read_only=True)
    pedidos: list[Pedido] = []

    for ws in wb.worksheets:
        def cel(key):
            r, c = _CEL[key]
            return ws.cell(r, c).value

        fecha = cel("fecha")
        ped = Pedido(
            hoja=ws.title.strip(),
            cliente=_txt(cel("cliente")),
            nro_cliente=_txt(cel("nro_cliente")),
            nro_vendedor=_txt(cel("nro_vendedor")),
            nro_pedido=_txt(cel("nro_pedido")),
            cond_pago=_txt(cel("cond_pago")),
            fecha=fecha if isinstance(fecha, datetime) else None,
            total_con_iva_declarado=_num(cel("total_con_iva")),
        )

        for fila, row in enumerate(
            ws.iter_rows(min_row=_FILA_ITEMS, values_only=True), start=_FILA_ITEMS
        ):
            def col(idx):  # acceso seguro a filas mas cortas que la grilla
                return row[idx - 1] if idx - 1 < len(row) else None

            cant = _num(col(_COL_CANT))
            if not cant or cant <= 0:
                continue  # relleno del catalogo o encabezado de rubro
            ped.items.append(
                Item(
                    fila=fila,
                    codigo=_txt(col(_COL_CODIGO)),
                    descripcion=_txt(col(_COL_DESC)),
                    und_x_caja=_txt(col(_COL_UXC)),
                    precio_sin_iva=_num(col(_COL_PRECIO)) or 0.0,
                    cantidad=cant,
                    subtotal=_num(col(_COL_SUBTOT)) or 0.0,
                )
            )
        # Saltar hojas que son sólo relleno del template: sin cliente,
        # sin Nro. Cliente y sin ítems con cantidad > 0. Aparecen porque
        # el Excel trae siempre plantilla fija con varias hojas vacías.
        if (
            not ped.cliente.strip()
            and not ped.nro_cliente.strip()
            and not ped.items
        ):
            continue
        pedidos.append(ped)

    wb.close()
    return pedidos


def _demo(path: str) -> None:
    pedidos = leer_pedidos(path)
    print(f"Archivo: {path}")
    print(f"Pedidos encontrados: {len(pedidos)}\n")
    for p in pedidos:
        sello = "OK " if p.total_ok else "REVISAR"
        combo = "  [tiene combo]" if p.tiene_combos else ""
        print(f"== {p.hoja} | {p.cliente} (Nro {p.nro_cliente}) | "
              f"vend {p.nro_vendedor} | pedido {p.nro_pedido}{combo}")
        if p.cond_pago:
            print(f"   Cond. de pago: {p.cond_pago}")
        for i in p.items:
            print(f"   {i.codigo:<16} {i.descripcion[:34]:<34} "
                  f"cant={i.cantidad:<6g} ${i.precio_sin_iva:>9.2f}  "
                  f"subt=${i.subtotal:>10.2f}")
        print(f"   --> suma sin IVA ${p.suma_subtotales:,.2f} | "
              f"con IVA calc ${p.total_con_iva_calculado:,.2f} | "
              f"declarado ${p.total_con_iva_declarado:,.2f} | "
              f"control: {sello}\n")


if __name__ == "__main__":
    import sys

    _demo(sys.argv[1] if len(sys.argv) > 1
          else "assets/Imágenes Proceso Carga de Pedido/pedidos de 14 mayo 26.xlsx")
