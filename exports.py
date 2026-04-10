"""
exports.py — Generación de archivos exportables para los vendedores.

Función principal: `exportar_agenda_vendedor()` que genera un .xlsx con
5 hojas listas para que cada vendedor se lleve "su agenda" después de
la reunión semanal:

  1. Resumen          — performance del período + cobertura + vs equipo
  2. Mi cartera       — listado completo con monto mes/semana, ¿compró?
  3. Clientes dormidos — los que no compraron este mes
  4. Penetración      — % de su cartera por sub-rubro
  5. Top 80%          — los pocos clientes que generan la mayor parte

Uso:
    buf = exporter.exportar_agenda_vendedor(df_sem, df_mes, df_clientes, v)
    # buf es un BytesIO, listo para st.download_button.

Funciones puras: no importa streamlit, no escribe a disco.
"""

import io
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

import metrics


# =====================================================================
# Estilos consistentes con el theme Dieter Rams del dashboard
# =====================================================================

INK = "1A1A1A"
TEXT_SOFT = "767676"
LINE = "E0E0E0"
ACCENT = "C8552F"

TITLE_FONT = Font(bold=True, size=14, color=INK)
SECTION_FONT = Font(bold=True, size=11, color=INK)
LABEL_FONT = Font(size=10, color=TEXT_SOFT)
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor=INK)
BOLD = Font(bold=True, size=10)

THIN_SIDE = Side(border_style="thin", color=LINE)
THIN_BORDER = Border(
    left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE
)

MONEY_FMT = '"$"#,##0'
INT_FMT = "#,##0"
PCT_FMT = "0.0%"

CENTER = Alignment(horizontal="center", vertical="center")


# =====================================================================
# Helpers internos
# =====================================================================

def _write_header_row(ws, row: int, columns: list) -> None:
    """Pinta una fila como header (negro con texto blanco)."""
    for col_idx, label in enumerate(columns, start=1):
        c = ws.cell(row=row, column=col_idx, value=label)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER


def _set_widths(ws, widths: list) -> None:
    """Setea anchos de columnas."""
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# =====================================================================
# Función principal
# =====================================================================

def exportar_agenda_vendedor(
    df_sem: pd.DataFrame,
    df_mes: pd.DataFrame,
    df_clientes: pd.DataFrame,
    vendedor: str,
) -> io.BytesIO:
    """
    Genera la agenda personal de un vendedor como xlsx con 5 hojas.

    Args:
        df_sem: facturación semanal preparada (post transforms).
        df_mes: facturación mensual preparada.
        df_clientes: maestro de clientes.
        vendedor: email del vendedor (debe estar en df_clientes.vendedor).

    Devuelve un `io.BytesIO` con el xlsx generado, listo para pasar
    a `st.download_button(data=...)`.
    """
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Resumen"
    _build_resumen(ws1, vendedor, df_sem, df_mes, df_clientes)

    ws2 = wb.create_sheet("Mi cartera")
    _build_mi_cartera(ws2, vendedor, df_sem, df_mes, df_clientes)

    ws3 = wb.create_sheet("Clientes dormidos")
    _build_clientes_dormidos(ws3, vendedor, df_mes, df_clientes)

    ws4 = wb.create_sheet("Penetración")
    _build_penetracion(ws4, vendedor, df_mes, df_clientes)

    ws5 = wb.create_sheet("Top 80%")
    _build_top_80(ws5, vendedor, df_mes, df_clientes)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# =====================================================================
# Hoja 1 — Resumen
# =====================================================================

def _build_resumen(ws, vendedor, df_sem, df_mes, df_clientes):
    """Big numbers + cobertura + comparación vs el promedio del equipo."""
    ws["A1"] = f"Agenda — {vendedor}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")

    ws["A2"] = f"Generada: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:D2")

    # Filtros para el vendedor
    df_v_mes = df_mes[df_mes["vendedor"] == vendedor]
    df_v_sem = df_sem[df_sem["vendedor"] == vendedor]

    monto_mes = float(df_v_mes["monto"].sum()) if not df_v_mes.empty else 0.0
    monto_sem = float(df_v_sem["monto"].sum()) if not df_v_sem.empty else 0.0
    unidades_mes = (
        int(df_v_mes["unidades"].sum()) if not df_v_mes.empty else 0
    )
    unidades_sem = (
        int(df_v_sem["unidades"].sum()) if not df_v_sem.empty else 0
    )

    # Sección 1: Performance
    row = 4
    ws.cell(row=row, column=1, value="PERFORMANCE DEL PERÍODO").font = SECTION_FONT
    row += 1

    pares = [
        ("Total mes", monto_mes, MONEY_FMT),
        ("Unidades mes", unidades_mes, INT_FMT),
        ("Total semana", monto_sem, MONEY_FMT),
        ("Unidades semana", unidades_sem, INT_FMT),
    ]
    for label, value, fmt in pares:
        ws.cell(row=row, column=1, value=label)
        c = ws.cell(row=row, column=2, value=value)
        c.number_format = fmt
        c.font = BOLD
        row += 1

    # Sección 2: Cobertura
    row += 1
    ws.cell(row=row, column=1, value="COBERTURA DE CARTERA").font = SECTION_FONT
    row += 1

    cob = metrics.cobertura_por_vendedor(df_mes, df_clientes)
    cob_v = cob[cob["vendedor"] == vendedor] if not cob.empty else pd.DataFrame()

    if not cob_v.empty:
        r = cob_v.iloc[0]
        ws.cell(row=row, column=1, value="Clientes asignados")
        ws.cell(row=row, column=2, value=int(r["clientes_asignados"]))
        row += 1
        ws.cell(row=row, column=1, value="Clientes con venta este mes")
        ws.cell(row=row, column=2, value=int(r["clientes_con_venta"]))
        row += 1
        ws.cell(row=row, column=1, value="% Cobertura")
        c = ws.cell(row=row, column=2, value=float(r["cobertura_pct"]) / 100)
        c.number_format = PCT_FMT
        c.font = BOLD
        row += 1
    else:
        ws.cell(row=row, column=1, value="(sin datos de cobertura)").font = LABEL_FONT
        row += 1

    # Sección 3: Comparación vs equipo
    row += 1
    ws.cell(row=row, column=1, value="VS PROMEDIO DEL EQUIPO").font = SECTION_FONT
    row += 1

    ventas_equipo = metrics.ventas_por_vendedor(df_mes)
    if not ventas_equipo.empty:
        promedio = float(ventas_equipo["monto_total"].mean())
        diff = monto_mes - promedio
        diff_pct = (diff / promedio) if promedio > 0 else 0

        ws.cell(row=row, column=1, value="Promedio mes del equipo")
        c = ws.cell(row=row, column=2, value=promedio)
        c.number_format = MONEY_FMT
        row += 1

        ws.cell(row=row, column=1, value="Diferencia vs promedio")
        c = ws.cell(row=row, column=2, value=diff)
        c.number_format = MONEY_FMT
        c.font = BOLD
        row += 1

        ws.cell(row=row, column=1, value="% sobre / bajo promedio")
        c = ws.cell(row=row, column=2, value=diff_pct)
        c.number_format = PCT_FMT
        c.font = BOLD

    _set_widths(ws, [32, 18, 4, 18])


# =====================================================================
# Hoja 2 — Mi cartera
# =====================================================================

def _build_mi_cartera(ws, vendedor, df_sem, df_mes, df_clientes):
    """Listado completo de la cartera con monto mes, semana, unidades."""
    ws["A1"] = "Mi cartera completa"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")

    cartera = (
        df_clientes[df_clientes["vendedor"] == vendedor]
        .dropna(subset=["documento"])
        .drop_duplicates(subset="documento")
        .copy()
    )

    if cartera.empty:
        ws["A3"] = "Sin clientes en cartera."
        return

    # Match estricto: solo FAC del propio vendedor a sus clientes
    df_v_mes = df_mes[
        (df_mes["vendedor"] == vendedor) & (df_mes["tipo"] == "FAC")
    ]
    df_v_sem = df_sem[
        (df_sem["vendedor"] == vendedor) & (df_sem["tipo"] == "FAC")
    ]

    monto_mes_map = (
        df_v_mes.groupby("documento")["monto"].sum().to_dict()
        if not df_v_mes.empty
        else {}
    )
    monto_sem_map = (
        df_v_sem.groupby("documento")["monto"].sum().to_dict()
        if not df_v_sem.empty
        else {}
    )
    unid_mes_map = (
        df_v_mes.groupby("documento")["unidades"].sum().to_dict()
        if not df_v_mes.empty
        else {}
    )

    cartera["monto_mes"] = cartera["documento"].map(monto_mes_map).fillna(0)
    cartera["monto_semana"] = (
        cartera["documento"].map(monto_sem_map).fillna(0)
    )
    cartera["unidades_mes"] = (
        cartera["documento"].map(unid_mes_map).fillna(0)
    )
    cartera["compro_mes"] = cartera["monto_mes"] > 0

    # Ordenar: los que compraron arriba (por monto desc), los dormidos abajo
    cartera = cartera.sort_values(
        ["compro_mes", "monto_mes"], ascending=[False, False]
    )

    headers = [
        "Documento",
        "Razón Social",
        "Monto mes",
        "Monto semana",
        "Unidades mes",
        "¿Compró este mes?",
    ]
    _write_header_row(ws, 3, headers)

    row = 4
    for _, r in cartera.iterrows():
        ws.cell(row=row, column=1, value=r["documento"])
        ws.cell(row=row, column=2, value=r["razon_social"])
        c = ws.cell(row=row, column=3, value=float(r["monto_mes"]))
        c.number_format = MONEY_FMT
        c = ws.cell(row=row, column=4, value=float(r["monto_semana"]))
        c.number_format = MONEY_FMT
        c = ws.cell(row=row, column=5, value=int(r["unidades_mes"]))
        c.number_format = INT_FMT
        ws.cell(row=row, column=6, value="Sí" if r["compro_mes"] else "NO")
        row += 1

    _set_widths(ws, [14, 42, 16, 16, 14, 18])


# =====================================================================
# Hoja 3 — Clientes dormidos
# =====================================================================

def _build_clientes_dormidos(ws, vendedor, df_mes, df_clientes):
    """Lista de clientes en cartera que no compraron este mes."""
    ws["A1"] = "Clientes dormidos (no compraron este mes)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")
    ws["A2"] = (
        "Match estricto: solo cuenta venta tipo FAC hecha por el propio "
        "vendedor asignado al cliente."
    )
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:C2")

    cartera = (
        df_clientes[df_clientes["vendedor"] == vendedor]
        .dropna(subset=["documento"])
        .drop_duplicates(subset="documento")
    )

    df_v_mes = df_mes[
        (df_mes["vendedor"] == vendedor) & (df_mes["tipo"] == "FAC")
    ]
    docs_con_venta = (
        set(df_v_mes["documento"].unique()) if not df_v_mes.empty else set()
    )

    dormidos = cartera[~cartera["documento"].isin(docs_con_venta)]

    if dormidos.empty:
        ws["A4"] = "Sin clientes dormidos — toda tu cartera compró este mes."
        ws["A4"].font = BOLD
        return

    dormidos = dormidos.sort_values("razon_social", na_position="last")

    headers = ["Documento", "Razón Social", "Estado"]
    _write_header_row(ws, 4, headers)

    row = 5
    for _, r in dormidos.iterrows():
        ws.cell(row=row, column=1, value=r["documento"])
        ws.cell(row=row, column=2, value=r["razon_social"])
        ws.cell(row=row, column=3, value="Sin compras este mes")
        row += 1

    _set_widths(ws, [14, 42, 26])


# =====================================================================
# Hoja 4 — Penetración por sub-rubro
# =====================================================================

def _build_penetracion(ws, vendedor, df_mes, df_clientes):
    """Solo la fila del vendedor de la matriz de penetración."""
    ws["A1"] = "Penetración por sub-rubro (mes)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")
    ws["A2"] = (
        "% de tu cartera asignada que recibió al menos una venta de "
        "cada sub-rubro este mes. Los % bajos son tus huecos de cross-sell."
    )
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:B2")

    pivot = metrics.penetracion_por_sub_rubro_pivot(df_mes, df_clientes)

    if pivot.empty or vendedor not in pivot.index:
        ws["A4"] = "Sin datos de penetración para este vendedor."
        return

    fila = pivot.loc[vendedor].sort_values(ascending=False)

    headers = ["Sub-rubro", "% Cobertura"]
    _write_header_row(ws, 4, headers)

    row = 5
    for sub_rubro, pct in fila.items():
        ws.cell(row=row, column=1, value=str(sub_rubro))
        c = ws.cell(row=row, column=2, value=float(pct) / 100)
        c.number_format = PCT_FMT
        row += 1

    _set_widths(ws, [25, 16])


# =====================================================================
# Hoja 5 — Top 80% de clientes
# =====================================================================

def _build_top_80(ws, vendedor, df_mes, df_clientes):
    """Los pocos clientes que generan ~80% de la venta del vendedor."""
    ws["A1"] = "Top 80% de tus clientes — los que hay que blindar"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:E1")
    ws["A2"] = (
        "Pareto 80/20: los pocos clientes que generan la mayor parte "
        "de tu venta. Antes de buscar clientes nuevos, asegurate de "
        "que estos no se vayan."
    )
    ws["A2"].font = LABEL_FONT
    ws.merge_cells("A2:E2")

    par = metrics.pareto_clientes(df_mes, df_clientes, vendedor=vendedor)

    if par.empty:
        ws["A4"] = "Sin datos de Pareto para este vendedor."
        return

    # Filtrar al CORE 80%. Si todos los clientes están concentrados en
    # pocos y el primer cliente ya supera el 80%, igual incluirlo.
    core = par[par["pct_acumulado"] <= 80.0]
    if core.empty:
        core = par.head(1)

    headers = [
        "Documento",
        "Razón Social",
        "Monto",
        "% Individual",
        "% Acumulado",
    ]
    _write_header_row(ws, 4, headers)

    row = 5
    for _, r in core.iterrows():
        ws.cell(row=row, column=1, value=r["documento"])
        ws.cell(row=row, column=2, value=r["razon_social"])
        c = ws.cell(row=row, column=3, value=float(r["monto"]))
        c.number_format = MONEY_FMT
        c = ws.cell(row=row, column=4, value=float(r["pct_individual"]) / 100)
        c.number_format = PCT_FMT
        c = ws.cell(row=row, column=5, value=float(r["pct_acumulado"]) / 100)
        c.number_format = PCT_FMT
        row += 1

    # Línea de cierre con totales
    row += 1
    ws.cell(row=row, column=1, value=f"TOTAL: {len(core)} clientes en CORE 80%").font = BOLD

    _set_widths(ws, [14, 42, 16, 14, 14])
